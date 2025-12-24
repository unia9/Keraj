# -*- coding: utf-8 -*-
import os
import sys
import json
import traceback
import threading
import datetime as dt
from pathlib import Path




import pandas as pd

# --- Tkinter / GUI ---
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    TkBase = TkinterDnD.Tk
    USE_DND = True
except Exception:
    import tkinter as tk
    TkBase = tk.Tk
    USE_DND = False

import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
from tkinter.simpledialog import askstring

try:
    import ttkbootstrap as tb
    from ttkbootstrap import Style as TtkbStyle
    USE_TTKB = True
except Exception:
    tb = None
    TtkbStyle = None
    USE_TTKB = False

# opcjonalny nowoczesny interfejs (CustomTkinter)
try:
    import customtkinter as ctk
    USE_CTK = True
except Exception:
    ctk = None
    USE_CTK = False

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

APP_TITLE = "Wyniki 5 – SP Górzno"
ARCHIVE_TITLE = "Wyniki 5 – Archiwum wyników"

# ---------- zasoby (PyInstaller) ----------
def resource_path(rel_path: str) -> Path:
    base = Path(getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__))))
    return base / rel_path


PATH_LOGO = str(resource_path("logo_szkoly.png"))
PATH_PDF = str(resource_path("Instrukcja_Wyniki5.pdf"))


# ---------- konfiguracja (APPDATA) ----------
def appdata_dir() -> Path:
    d = Path(os.getenv("APPDATA", Path.home())) / "Wyniki5"
    d.mkdir(parents=True, exist_ok=True)
    return d


def cfg_path() -> Path:
    return appdata_dir() / "config.json"


def load_cfg() -> dict:
    p = cfg_path()
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_cfg(cfg: dict):
    cfg_path().write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")

# ---------- konfiguracja zakładek przedmiotów w archiwum ----------
def subject_tabs_path() -> Path:
    """Zwraca ścieżkę do pliku z konfiguracją zakładek przedmiotów w APPDATA."""
    d = appdata_dir()
    d.mkdir(parents=True, exist_ok=True)
    return d / "subject_tabs.json"


def load_subject_tabs_config() -> dict:
    """Wczytuje konfigurację zakładek przedmiotów z pliku JSON.

    Struktura:
    {
        "tabs": {
            "Historia": {"label": "Historia", "visible": true},
            "WOS": {"label": "Wiedza o społeczeństwie", "visible": true},
            ...
        }
    }
    """
    p = subject_tabs_path()
    if p.exists():
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                tabs = data.get("tabs")
                if isinstance(tabs, dict):
                    return data
                # jeśli stary format (słownik bez klucza "tabs") – opakuj
                return {"tabs": data}
        except Exception:
            return {"tabs": {}}
    return {"tabs": {}}


def save_subject_tabs_config(cfg: dict) -> None:
    """Zapisuje konfigurację zakładek przedmiotów do pliku JSON."""
    p = subject_tabs_path()
    try:
        p.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        # brak twardego błędu – ustawienia zakładek są opcjonalne
        pass




# ---------- archiwum wyników (JSON) ----------
def _archive_dir() -> Path:
    d = appdata_dir() / "archiwum"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _slugify(text: str) -> str:
    allowed = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_"
    slug = []
    for ch in text:
        if ch in allowed:
            slug.append(ch)
        elif ch.isspace() or ch in ".,;:/\\|":
            slug.append("_")
        else:
            continue
    cleaned = "".join(slug).strip("_")
    return cleaned or "wyniki"


def save_result_to_archive(context_name: str, title: str, df: pd.DataFrame, meta: dict | None = None) -> Path:
    r"""
    Zapisuje wyniki pojedynczego sprawdzianu do archiwum (JSON) w APPDATA\Wyniki5\archiwum.

    context_name : nazwa kontekstu / szkoły (np. „SP Górzno”)
    title        : opis sprawdzianu (np. „6A Historia – Sprawdzian 1”)
    df           : DataFrame z wynikami (po przeliczeniu)
    meta         : dodatkowe informacje (max_points, plik źródłowy, metoda, itp.)
    """
    arch_dir = _archive_dir()
    if meta is None:
        meta = {}

    created = dt.datetime.now().isoformat(timespec="seconds")

    df_for_json = df.copy()
    df_for_json = df_for_json.astype(object).where(pd.notnull(df_for_json), None)

    payload = {
        "schema": 1,
        "context": context_name,
        "title": title,
        "created": created,
        "meta": meta,
        "columns": [str(c) for c in df_for_json.columns],
        "rows": df_for_json.values.tolist(),
    }

    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_title = _slugify(title)[:40]
    filename = f"{ts}_{safe_title}.json"
    out_path = arch_dir / filename
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return out_path


# ======================================================================
# ARCHIWUM – podgląd tabeli w osobnym oknie
# ======================================================================


class StudentHistoryWindow(tk.Toplevel):
    """
    Okno z przekrojowym raportem „Historia ucznia” na podstawie archiwum.
    """

    def __init__(self, master, student_name: str, records: list[dict]):
        super().__init__(master)
        self.title(f"Historia ucznia – {student_name}")
        self.minsize(900, 400)
        self.transient(master)
        self.grab_set()

        self._student_name = student_name
        self._df = pd.DataFrame(records) if records else pd.DataFrame(
            columns=["Data", "Kontekst", "Klasa", "Przedmiot", "Punkty", "Procent", "Ocena"]
        )

        pad = 8
        main = ttk.Frame(self, padding=pad)
        main.pack(fill="both", expand=True)

        ttk.Label(main, text=f"Historia ucznia: {student_name}", font=("TkDefaultFont", 10, "bold")).pack(
            anchor="w", pady=(0, 4)
        )

        columns = ["Data", "Kontekst", "Klasa", "Przedmiot", "Punkty", "Procent", "Ocena"]
        self.tree = ttk.Treeview(main, columns=columns, show="headings", height=12)
        
        vsb = ttk.Scrollbar(main, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)

        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Najpierw wstaw dane, aby móc obliczyć szerokości na podstawie zawartości
        all_values = []
        if not self._df.empty:
            for _, row in self._df.iterrows():
                values = [
                    str(row.get("Data", "")),
                    str(row.get("Kontekst", "")),
                    str(row.get("Klasa", "")),
                    str(row.get("Przedmiot", "")),
                    str(row.get("Punkty", "")),
                    str(row.get("Procent", "")),
                    str(row.get("Ocena", "")),
                ]
                all_values.append(values)
                self.tree.insert("", "end", values=values)

        # Oblicz szerokości kolumn na podstawie zawartości (nazwa kolumny + najdłuższa wartość)
        for col_idx, col in enumerate(columns):
            self.tree.heading(col, text=col)
            anchor = "center" if col in ("Data", "Punkty", "Procent", "Ocena") else "w"
            # Szerokość na podstawie nazwy kolumny
            header_width = len(col) * 8
            # Szerokość na podstawie najdłuższej wartości w kolumnie
            max_value_width = 0
            if all_values:
                for values in all_values:
                    if col_idx < len(values):
                        val_str = str(values[col_idx])
                        max_value_width = max(max_value_width, len(val_str) * 7)
            # Użyj większej z dwóch wartości, z minimalną szerokością 80 i maksymalną 300
            width = min(max(80, max(header_width, max_value_width + 10)), 300)
            self.tree.column(col, width=width, anchor=anchor)

        bottom = ttk.Frame(self, padding=(pad, 0, pad, pad))
        bottom.pack(fill="x")

        ttk.Button(bottom, text="Eksport do Excela…", command=self._export_excel).pack(side="left")
        ttk.Button(bottom, text="Eksport do PDF…", command=self._export_pdf).pack(side="left", padx=(8, 0))

        ttk.Button(bottom, text="Zamknij", command=self.destroy).pack(side="right")

    def _export_excel(self):
        if self._df is None or self._df.empty:
            messagebox.showinfo("Historia ucznia", "Brak danych do eksportu.")
            return

        from tkinter import filedialog as _filedialog

        path = _filedialog.asksaveasfilename(
            title="Zapisz historię ucznia",
            defaultextension=".xlsx",
            filetypes=[("Plik Excel", "*.xlsx"), ("Wszystkie pliki", "*.*")],
        )
        if not path:
            return

        try:
            self._df.to_excel(path, index=False)
        except Exception as e:
            messagebox.showerror(
                "Historia ucznia",
                f"Nie udało się zapisać pliku:\n{e}",
            )
            return

        messagebox.showinfo(
            "Historia ucznia",
            "Pomyślnie zapisano plik Excela.\nPDF możesz łatwo utworzyć z poziomu Excela (Zapisz jako PDF).",
        )

    def _export_pdf(self):
        """
        Eksportuje historię ucznia do prostego raportu PDF.
        Wymaga zainstalowanej biblioteki reportlab: pip install reportlab
        """
        if self._df is None or self._df.empty:
            messagebox.showinfo("Historia ucznia", "Brak danych do eksportu.")
            return

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except Exception:
            messagebox.showerror(
                "Historia ucznia",
                "Do eksportu PDF wymagana jest biblioteka 'reportlab'.\n"
                "Zainstaluj ją poleceniem:\n"
                "pip install reportlab",
            )
            return

        from tkinter import filedialog as _filedialog

        path = _filedialog.asksaveasfilename(
            title="Zapisz historię ucznia (PDF)",
            defaultextension=".pdf",
            filetypes=[("Plik PDF", "*.pdf"), ("Wszystkie pliki", "*.*")],
        )
        if not path:
            return

        try:
            # konfiguracja fontów z obsługą polskich znaków
            font_name = "Helvetica"
            try:
                pdfmetrics.registerFont(TTFont("DejaVuSans", "DejaVuSans.ttf"))
                font_name = "DejaVuSans"
            except Exception:
                try:
                    pdfmetrics.registerFont(TTFont("Arial", "C:/Windows/Fonts/arial.ttf"))
                    font_name = "Arial"
                except Exception:
                    font_name = "Helvetica"

            c = canvas.Canvas(path, pagesize=A4)
            width, height = A4
            left_margin = 40
            top_margin = height - 40
            line_height = 14

            c.setFont(font_name, 14)
            c.drawString(left_margin, top_margin, f"Historia ucznia: {self._student_name}")
            y = top_margin - 2 * line_height

            c.setFont(font_name, 9)
            header = "Data | Kontekst | Klasa | Przedmiot | Punkty | Procent | Ocena"
            c.drawString(left_margin, y, header)
            y -= line_height
            c.line(left_margin, y + 3, width - left_margin, y + 3)
            y -= line_height

            for _, row in self._df.iterrows():
                if y < 40:
                    c.showPage()
                    c.setFont(font_name, 14)
                    c.drawString(left_margin, top_margin, f"Historia ucznia: {self._student_name}")
                    y = top_margin - 2 * line_height
                    c.setFont(font_name, 9)
                data_str = (
                    f"{row.get('Data', '')} | "
                    f"{row.get('Kontekst', '')} | "
                    f"{row.get('Klasa', '')} | "
                    f"{row.get('Przedmiot', '')} | "
                    f"{row.get('Punkty', '')} | "
                    f"{row.get('Procent', '')} | "
                    f"{row.get('Ocena', '')}"
                )
                c.drawString(left_margin, y, data_str[:2000])
                y -= line_height

            c.showPage()
            c.save()
        except Exception as e:
            messagebox.showerror(
                "Historia ucznia",
                f"Nie udało się zapisać pliku PDF:\n{e}",
            )
            return

        messagebox.showinfo("Historia ucznia", "Pomyślnie zapisano raport PDF.")

class ArchiveViewer(tk.Toplevel):
    """
    Okno podglądu archiwum wyników.
    """

    def __init__(self, master=None, preselect: Path | None = None):
        super().__init__(master)
        self.title(ARCHIVE_TITLE)
        self.minsize(900, 500)
        self.configure(bg="#F4F6FB")
        self.transient(master)
        self.grab_set()

        self._items_index: dict[str, Path] = {}
        self._current_df: pd.DataFrame | None = None
        self._current_meta: dict | None = None
        self._preselect_path: Path | None = Path(preselect) if preselect else None
        self._all_records: list[dict] = []
        self._filter_text = tk.StringVar()
        self._student_filter: str | None = None
        # filtr po przedmiocie w archiwum
        self._subject_filter = tk.StringVar()

        # filtr po roku szkolnym w archiwum (np. 2024/2025)
        self._school_year_var = tk.StringVar()

        # zakładki z przedmiotami (Notebook)
        self._subject_notebook: ttk.Notebook | None = None

        # konfiguracja zakładek przedmiotów (label/visible)
        self._subject_tabs_cfg: dict = load_subject_tabs_config()

        # stan sortowania kolumn w tabeli archiwum
        self._sort_state: dict[str, bool] = {}

        self._build_ui()
        self._refresh_list()

    def _build_ui(self):
        pad = 10

        # Helper: tworzy kolorową ramkę wokół LabelFrame w archiwum
        def create_colored_section(text: str, color: str, parent=None, **lf_kwargs):
            p = parent if parent is not None else self
            outer = tk.Frame(p, bg=color, highlightthickness=3, highlightbackground=color, bd=0)
            outer.pack(fill="x", padx=pad, pady=(0, pad))
            lf = ttk.LabelFrame(outer, text=text, padding=pad, style="Flat.TLabelframe", **lf_kwargs)
            lf.pack(fill="both", expand=True, padx=6, pady=6)
            return lf

        top = ttk.Frame(self, padding=pad)
        top.pack(fill="x")
        ttk.Label(
            top,
            text=(
                "Archiwum wyników programu „Wyniki 5”.\n"
                "Po lewej wybierz sprawdzian, po prawej zobaczysz tabelę z wynikami."
            ),
            justify="left",
        ).pack(anchor="w")

        main = ttk.Frame(self, padding=(pad, 0, pad, pad))
        main.pack(fill="both", expand=True)

        # lewa strona – lista zapisów
        left = ttk.Frame(main)
        left.pack(side="left", fill="y")

        ttk.Label(left, text="Zapisane sprawdziany:").pack(anchor="w")

        # Pole filtra – wyszukiwanie po szkole, klasie, tytule i uczniu
        filter_frame = ttk.Frame(left)
        filter_frame.pack(fill="x", pady=(2, 4))
        ttk.Label(filter_frame, text="Filtruj (uczeń / tytuł / szkoła):", style="Flat.TLabel").pack(side="left")
        entry_filter = ttk.Entry(filter_frame, textvariable=self._filter_text, width=30, style="Flat.TEntry")
        entry_filter.pack(side="left", fill="x", expand=True, padx=(4, 0))
        self._filter_text.trace_add("write", lambda *args: self._apply_filter())

        # Pole wyboru roku szkolnego (np. 2024/2025) – filtruje listę zapisów
        year_frame = ttk.Frame(left)
        year_frame.pack(fill="x", pady=(0, 4))
        ttk.Label(year_frame, text="Rok szkolny:", style="Flat.TLabel").pack(side="left")
        self._year_combo = ttk.Combobox(
            year_frame,
            textvariable=self._school_year_var,
            state="readonly",
            width=12,
        )
        self._year_combo.pack(side="left", padx=(4, 0))
        # zmiana roku szkolnego powoduje przebudowanie listy
        self._school_year_var.trace_add("write", lambda *args: self._rebuild_treeview())


        # zakładki z przedmiotami (Notebook) – wizualny podział
        self._subject_notebook = ttk.Notebook(left)
        self._subject_notebook.pack(fill="x", pady=(0, 4))
        # na start jedna zakładka „Wszystkie” – reszta zostanie zbudowana po wczytaniu archiwum
        frame_all = ttk.Frame(self._subject_notebook)
        self._subject_notebook.add(frame_all, text="Wszystkie")
        self._subject_notebook.bind("<<NotebookTabChanged>>", self._on_subject_tab_changed)

        cols = ("data", "context", "class", "subject", "school", "title")
        self.tree_tests = ttk.Treeview(left, columns=cols, show="headings", height=18)
        self.tree_tests.heading("data", text="Data")
        self.tree_tests.heading("context", text="Kontekst")
        self.tree_tests.heading("class", text="Klasa / grupa")
        self.tree_tests.heading("subject", text="Przedmiot")
        self.tree_tests.heading("school", text="Szkoła")
        self.tree_tests.heading("title", text="Opis sprawdzianu")

        # umożliwienie sortowania po kliknięciu w nagłówek kolumny
        self.tree_tests.heading("data", text="Data", command=lambda: self._on_tree_heading_click("data"))
        self.tree_tests.heading("context", text="Kontekst", command=lambda: self._on_tree_heading_click("context"))
        self.tree_tests.heading("class", text="Klasa / grupa", command=lambda: self._on_tree_heading_click("class"))
        self.tree_tests.heading("subject", text="Przedmiot", command=lambda: self._on_tree_heading_click("subject"))
        self.tree_tests.heading("school", text="Szkoła", command=lambda: self._on_tree_heading_click("school"))
        self.tree_tests.heading("title", text="Opis sprawdzianu", command=lambda: self._on_tree_heading_click("title"))
        self.tree_tests.column("data", width=90, anchor="w")
        self.tree_tests.column("context", width=120, anchor="w")
        self.tree_tests.column("class", width=90, anchor="w")
        self.tree_tests.column("subject", width=100, anchor="w")
        self.tree_tests.column("school", width=140, anchor="w")
        self.tree_tests.column("title", width=240, anchor="w")

        self.tree_tests.pack(side="left", fill="y", expand=False)

        self.tree_tests.bind("<<TreeviewSelect>>", self._on_select_item)

        # prawa strona – tabela wyników + informacje o ocenianiu
        right = ttk.Frame(main)
        right.pack(side="left", fill="both", expand=True, padx=(pad, 0))

        ttk.Label(right, text="Tabela wyników:").pack(anchor="w", pady=(0, 4))

        # Opakuj tabelę w Frame z paddingiem, aby nagłówek nie przykrywał pierwszego wiersza
        table_frame = ttk.Frame(right)
        table_frame.pack(side="top", fill="both", expand=True, padx=0, pady=0)
        
        self.table = ttk.Treeview(table_frame, show="headings")

        self.table.pack(side="top", fill="both", expand=True, padx=0, pady=(2, 0))

        # przycisk: pokaż wszystkie sprawdziany danego ucznia
        btn_student_frame = ttk.Frame(right)
        btn_student_frame.pack(fill="x", pady=(4, 0))
        ttk.Button(
            btn_student_frame,
            text="Pokaż wszystkie sprawdziany tego ucznia",
            command=self._filter_by_selected_student,
            style="TButton",
        ).pack(side="left", padx=(0, 4))
        ttk.Button(
            btn_student_frame,
            text="Historia ucznia (raport)",
            command=self._show_student_history,
            style="TButton",
        ).pack(side="left", padx=(0, 4))

        # blok: informacje o ocenianiu
        info_frame = create_colored_section("Informacje o ocenianiu", "#E7D9FF", parent=right)

        self.info_max_label = ttk.Label(info_frame, text="Maksymalna liczba punktów (testu): –", style="Flat.TLabel")
        self.info_max_label.grid(row=0, column=0, sticky="w", padx=4, pady=(4, 2))

        self.info_method_label = ttk.Label(info_frame, text="Metoda oceniania: –", style="Flat.TLabel")
        self.info_method_label.grid(row=1, column=0, sticky="w", padx=4, pady=(0, 2))

        self.info_weight_label = ttk.Label(info_frame, text="Waga arkusza (jeśli użyto średniej ważonej): –", style="Flat.TLabel")
        self.info_weight_label.grid(row=2, column=0, sticky="w", padx=4, pady=(0, 2))

        self.info_summary_label = ttk.Label(info_frame, text="Podsumowanie wyników: –", style="Flat.TLabel")
        self.info_summary_label.grid(row=3, column=0, sticky="w", padx=4, pady=(0, 4))

        ttk.Label(info_frame, text="Kryteria ocen (progi procentowe):", style="Flat.TLabel").grid(
            row=4, column=0, sticky="w", padx=4
        )

        self.scale_table = ttk.Treeview(
            info_frame,
            columns=("from", "to", "label"),
            show="headings",
            height=5,
        )
        self.scale_table.heading("from", text="Od (%)")
        self.scale_table.heading("to", text="Do (%)")
        self.scale_table.heading("label", text="Ocena")

        self.scale_table.column("from", width=70, anchor="center")
        self.scale_table.column("to", width=70, anchor="center")
        self.scale_table.column("label", width=180, anchor="w")

        self.scale_table.grid(row=5, column=0, sticky="nsew", padx=4, pady=(2, 4))

        info_frame.columnconfigure(0, weight=1)

        # kolory wierszy – konfiguracja tagów dla ocen
        self._grade_colors = {
            "6": "#C6EFCE",
            "5": "#CCFFCC",
            "4": "#FFF2CC",
            "3": "#FFD966",
            "2": "#F4CCCC",
            "1": "#EA9999",
        }
        for g, color in self._grade_colors.items():
            self.table.tag_configure(f"grade_{g}", background=color, foreground="#000000")

        # dolny pasek
        bottom = ttk.Frame(self, padding=(pad, 0, pad, pad))
        bottom.pack(fill="x")

        ttk.Button(bottom, text="Odśwież listę", command=self._refresh_list).pack(side="left")
        ttk.Button(bottom, text="Eksportuj zaznaczony do Excela…", command=self._export_selected).pack(
            side="left", padx=(8, 0)
        )
        ttk.Button(bottom, text="Usuń zaznaczony z archiwum", command=self._delete_selected).pack(
            side="left", padx=(8, 0)
        )
        ttk.Button(bottom, text="Zakładki przedmiotów…", command=self._open_subject_tabs_settings).pack(
            side="left", padx=(8, 0)
        )
        ttk.Button(bottom, text="Zamknij", command=self.destroy).pack(side="right")


    def _on_tree_heading_click(self, col: str) -> None:
        """Obsługa kliknięcia w nagłówek kolumny – sortowanie wierszy w drzewie.

        Kliknięcie w nagłówek:
        - sortuje po danej kolumnie,
        - przełącza kierunek (rosnąco/malejąco),
        - aktualizuje etykiety nagłówków, dodając strzałkę ▲/▼ przy aktywnej kolumnie.
        """
        # aktualny stan sortowania dla kolumny (True = malejąco, False = rosnąco)
        current = self._sort_state.get(col, False)
        new_reverse = not current

        # wyczyść inne kolumny – sortujemy tylko po jednej na raz
        self._sort_state = {col: new_reverse}

        # wykonaj sortowanie
        self._sort_tree_column(col, reverse=new_reverse)

        # zaktualizuj nagłówki z uwzględnieniem strzałek
        labels = {
            "data": "Data",
            "context": "Kontekst",
            "class": "Klasa / grupa",
            "subject": "Przedmiot",
            "school": "Szkoła",
            "title": "Opis sprawdzianu",
        }
        for cid, base in labels.items():
            text = base
            if cid == col:
                text += " ▼" if new_reverse else " ▲"
            self.tree_tests.heading(cid, text=text)

    def _sort_tree_column(self, col: str, reverse: bool = False) -> None:
        """Sortuje wiersze w tabeli archiwum według wartości w danej kolumnie.

        Sortowanie odbywa się tylko na poziomie widoku (Treeview) – nie zmienia kolejności
        danych w archiwum, więc nie wpływa na inne funkcje.
        """
        tree = self.tree_tests
        # pobierz dane z widocznych wierszy
        rows = []
        for iid in tree.get_children(""):
            value = tree.set(iid, col)
            rows.append((value, iid))

        def _to_key(val: str):
            # najpierw spróbuj potraktować jako liczbę
            try:
                # zamiana przecinka na kropkę (np. 32,5)
                return float(str(val).replace(",", ".").strip())
            except Exception:
                pass
            # w pozostałych przypadkach sortujemy tekstowo, case-insensitive
            return str(val).lower()

        rows.sort(key=lambda x: _to_key(x[0]), reverse=reverse)

        # przestaw wiersze w drzewie w nowej kolejności
        for index, (_, iid) in enumerate(rows):
            tree.move(iid, "", index)
    def _refresh_list(self):
        arch_dir = _archive_dir()
        arch_dir.mkdir(parents=True, exist_ok=True)
        files = sorted(
            arch_dir.glob("*.json"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )

        # pełna lista rekordów do filtrowania
        self._all_records = []

        for path in files:
            # data modyfikacji pliku – przyda się do określenia roku szkolnego
            dt_mod = dt.datetime.fromtimestamp(path.stat().st_mtime)
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                continue

            created = data.get("created", "") or ""
            if "T" in created:
                created_disp = created.replace("T", " ")[:16]
            else:
                created_disp = created[:16]

            context = data.get("context", "") or ""
            meta = data.get("meta") or {}
            class_name = ""
            subject = ""
            school = ""
            if isinstance(meta, dict):
                class_name = str(meta.get("class_name", "") or "")
                subject = str(meta.get("subject", "") or "")
                school = str(meta.get("school", "") or "")
            title = data.get("title", path.stem)

            # zbuduj listę uczniów (z kolumny Nazwisko / Imię i nazwisko)
            cols = data.get("columns") or []
            rows = data.get("rows") or []
            students = ""
            try:
                idx_name = None
                for i, col in enumerate(cols):
                    if not isinstance(col, str):
                        continue
                    col_l = col.lower().strip()
                    if col_l in ("nazwisko", "imię i nazwisko", "imie i nazwisko"):
                        idx_name = i
                        break
                    if "nazwisk" in col_l or ("imi" in col_l and "nazw" in col_l) or "uczeń" in col_l or "uczen" in col_l:
                        idx_name = i
                        break
                if idx_name is not None:
                    names_list: list[str] = []
                    for row in rows:
                        if idx_name < len(row):
                            val = row[idx_name]
                            if val is None:
                                continue
                            sval = str(val).strip()
                            if sval:
                                names_list.append(sval)
                    students = " ".join(names_list)
            except Exception:
                students = ""

            # wyznaczenie roku szkolnego na podstawie daty utworzenia
            school_year = ""
            try:
                src = created or ""
                year = month = day = None
                if src:
                    parts = src[:10].split("-")
                    if len(parts) == 3:
                        year = int(parts[0])
                        month = int(parts[1])
                        day = int(parts[2])
                if year is None:
                    dt_obj = dt_mod
                    year, month, day = dt_obj.year, dt_obj.month, dt_obj.day
                if month >= 9:
                    school_year = f"{year}/{year + 1}"
                else:
                    school_year = f"{year - 1}/{year}"
            except Exception:
                school_year = ""

            self._all_records.append(
                {
                    "created": created_disp,
                    "context": context,
                    "class_name": class_name,
                    "subject": subject,
                    "school": school,
                    "title": title,
                    "path": path,
                    "students": students,
                    "school_year": school_year,
                }
            )

        
        # zaktualizuj listę przedmiotów oraz zakładki (Notebook)
        subjects = sorted(
            {
                (rec.get("subject") or "").strip()
                for rec in self._all_records
                if rec.get("subject") and str(rec.get("subject")).strip()
            }
        )

        # aktualizacja konfiguracji zakładek (dodaj nowe przedmioty, zachowaj etykiety i widoczność)
        cfg = self._subject_tabs_cfg or {}
        tabs_cfg = cfg.get("tabs") or {}

        for subj in subjects:
            if subj not in tabs_cfg:
                tabs_cfg[subj] = {"label": subj, "visible": True}

        cfg["tabs"] = tabs_cfg
        self._subject_tabs_cfg = cfg
        save_subject_tabs_config(cfg)

        # odbuduj zakładki w Notebooku na podstawie konfiguracji
        nb = getattr(self, "_subject_notebook", None)
        if nb is not None:
            for tab_id in nb.tabs():
                nb.forget(tab_id)

            # zawsze zakładka „Wszystkie”
            frame_all = ttk.Frame(nb)
            nb.add(frame_all, text="Wszystkie")

            # zakładki dla przedmiotów oznaczonych jako widoczne
            for subj in subjects:
                info = tabs_cfg.get(subj, {"label": subj, "visible": True})
                if not info.get("visible", True):
                    continue
                label = (info.get("label") or subj).strip() or subj
                frame = ttk.Frame(nb)
                nb.add(frame, text=label)

            # wybór zakładki zgodnej z aktualnym filtrem (subject key), jeśli możliwe
            try:
                current_subj_key = (self._subject_filter.get() or "").strip()
            except Exception:
                current_subj_key = ""

            target_label = "Wszystkie"
            if current_subj_key:
                info = tabs_cfg.get(current_subj_key)
                if info and info.get("visible", True):
                    target_label = (info.get("label") or current_subj_key).strip() or current_subj_key

            for tab_id in nb.tabs():
                try:
                    if nb.tab(tab_id, "text") == target_label:
                        nb.select(tab_id)
                        break
                except Exception:
                    continue

        # zaktualizuj listę lat szkolnych w comboboxie
        years = sorted(
            {rec.get("school_year") for rec in self._all_records if rec.get("school_year")},
            reverse=True,
        )
        if years:
            values = ["Wszystkie lata"] + years
            try:
                self._year_combo["values"] = values
            except Exception:
                pass
            try:
                current = self._school_year_var.get() or ""
            except Exception:
                current = ""
            if not current or current not in values:
                self._school_year_var.set("Wszystkie lata")

        # wyczyść bieżącą tabelę i przebuduj listę z użyciem aktualnego filtra
        self._set_table(None, None)
        self._rebuild_treeview()



    def _open_subject_tabs_settings(self):
        """
        Okno zarządzania zakładkami przedmiotów:
        - zmiana etykiety,
        - ukrywanie/pokazywanie zakładek,
        - dodawanie/usuwanie wpisów.
        """
        cfg = self._subject_tabs_cfg or {}
        tabs_cfg = cfg.get("tabs") or {}

        win = tk.Toplevel(self)
        win.title("Zakładki przedmiotów w archiwum")
        win.transient(self)
        win.grab_set()
        win.minsize(480, 260)

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill="both", expand=True)

        cols = ("subject", "label", "visible")
        tree = ttk.Treeview(frame, columns=cols, show="headings", height=10)
        tree.heading("subject", text="Klucz przedmiotu")
        tree.heading("label", text="Etykieta zakładki")
        tree.heading("visible", text="Widoczna")
        tree.column("subject", width=150, anchor="w")
        tree.column("label", width=200, anchor="w")
        tree.column("visible", width=80, anchor="center")

        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)

        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        def refresh_tree():
            tree.delete(*tree.get_children())
            cfg_local = self._subject_tabs_cfg or {}
            tabs_local = cfg_local.get("tabs") or {}
            for subj, info in sorted(tabs_local.items(), key=lambda x: (str(x[0]).lower())):
                if isinstance(info, dict):
                    label = (info.get("label") or subj).strip() or subj
                    vis = info.get("visible", True)
                else:
                    label = str(info).strip() or subj
                    vis = True
                tree.insert("", "end", iid=subj, values=(subj, label, "tak" if vis else "nie"))

        def on_add():
            from tkinter.simpledialog import askstring as _askstring
            subj = _askstring("Nowy przedmiot", "Podaj nazwę/klucz przedmiotu:")
            if not subj:
                return
            subj = subj.strip()
            if not subj:
                return
            cfg_local = self._subject_tabs_cfg or {}
            tabs_local = cfg_local.get("tabs") or {}
            if subj in tabs_local:
                messagebox.showinfo("Zakładki przedmiotów", "Taki przedmiot już istnieje w konfiguracji.")
                return
            tabs_local[subj] = {"label": subj, "visible": True}
            cfg_local["tabs"] = tabs_local
            self._subject_tabs_cfg = cfg_local
            save_subject_tabs_config(cfg_local)
            refresh_tree()
            self._refresh_list()

        def on_delete():
            sel = tree.selection()
            if not sel:
                return
            subj = sel[0]
            if not messagebox.askyesno(
                "Zakładki przedmiotów",
                f"Czy na pewno usunąć konfigurację przedmiotu: {subj}?",
            ):
                return
            cfg_local = self._subject_tabs_cfg or {}
            tabs_local = cfg_local.get("tabs") or {}
            if subj in tabs_local:
                tabs_local.pop(subj, None)
                cfg_local["tabs"] = tabs_local
                self._subject_tabs_cfg = cfg_local
                save_subject_tabs_config(cfg_local)
                refresh_tree()
                # jeśli aktualny filtr dotyczy usuniętego przedmiotu – wyczyść
                try:
                    if (self._subject_filter.get() or "").strip() == subj:
                        self._subject_filter.set("")
                except Exception:
                    pass
                self._refresh_list()

        def on_toggle_visible():
            sel = tree.selection()
            if not sel:
                return
            subj = sel[0]
            cfg_local = self._subject_tabs_cfg or {}
            tabs_local = cfg_local.get("tabs") or {}
            info = tabs_local.get(subj) or {}
            if not isinstance(info, dict):
                info = {"label": str(info), "visible": True}
            info["visible"] = not info.get("visible", True)
            tabs_local[subj] = info
            cfg_local["tabs"] = tabs_local
            self._subject_tabs_cfg = cfg_local
            save_subject_tabs_config(cfg_local)
            refresh_tree()
            self._refresh_list()

        def on_edit_label(event=None):
            sel = tree.selection()
            if not sel:
                return
            subj = sel[0]
            cfg_local = self._subject_tabs_cfg or {}
            tabs_local = cfg_local.get("tabs") or {}
            info = tabs_local.get(subj) or {}
            current_label = ""
            if isinstance(info, dict):
                current_label = (info.get("label") or subj).strip()
            else:
                current_label = str(info).strip() or subj

            from tkinter.simpledialog import askstring as _askstring
            new_label = _askstring("Edycja etykiety", f"Nowa etykieta dla przedmiotu '{subj}':", initialvalue=current_label)
            if not new_label:
                return
            new_label = new_label.strip()
            if not new_label:
                return

            if not isinstance(info, dict):
                info = {"label": current_label or subj, "visible": True}
            info["label"] = new_label
            tabs_local[subj] = info
            cfg_local["tabs"] = tabs_local
            self._subject_tabs_cfg = cfg_local
            save_subject_tabs_config(cfg_local)
            refresh_tree()
            self._refresh_list()

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill="x", pady=(8, 0))

        ttk.Button(btn_frame, text="Dodaj…", command=on_add).pack(side="left")
        ttk.Button(btn_frame, text="Usuń", command=on_delete).pack(side="left", padx=(8, 0))
        ttk.Button(btn_frame, text="Pokaż/ukryj", command=on_toggle_visible).pack(side="left", padx=(8, 0))
        ttk.Button(btn_frame, text="Zmień etykietę", command=on_edit_label).pack(side="left", padx=(8, 0))
        ttk.Button(btn_frame, text="Zamknij", command=win.destroy).pack(side="right")

        tree.bind("<Double-1>", on_edit_label)

        refresh_tree()
    def _apply_filter(self):
        """
        Reaguje na zmiany w polu „Filtruj…” i przebudowuje widok listy.

        Dodatkowo:
        - jeśli w polu filtra wpisano nazwisko / imię i nazwisko,
          po lewej filtruje archiwum,
          a po prawej próbuje zbudować przegląd wszystkich sprawdzianów tego ucznia.
        """
        # zapamiętujemy aktualny tekst filtra także jako potencjalny filtr ucznia
        try:
            text = (self._filter_text.get() or "").strip()
        except Exception:
            text = ""
        self._student_filter = text or None

        # najpierw przebuduj listę po lewej
        self._rebuild_treeview()

        # jeśli coś wpisano – spróbuj zbudować przegląd sprawdzianów tego ucznia po prawej
        if text:
            if not self._show_student_overview_for_filter(text):
                # jeśli nie udało się znaleźć ucznia – czyścimy filtr ucznia
                self._student_filter = None

    def _rebuild_treeview(self):
        """
        Buduje widok drzewa na podstawie self._all_records i tekstu filtra.

        Filtr działa na:
        - dacie,
        - kontekście (szkoła/placówka),
        - klasie / grupie,
        - tytule,
        - ORAZ na nazwiskach uczniów zapisanych w archiwum.
        """
        # wyczyść drzewko
        for iid in self.tree_tests.get_children():
            self.tree_tests.delete(iid)
        self._items_index.clear()

        try:
            filter_text = (self._filter_text.get() or "").strip().lower()
        except Exception:
            filter_text = ""

        try:
            year_filter = (self._school_year_var.get() or "").strip()
        except Exception:
            year_filter = ""

        try:
            subject_filter = (self._subject_filter.get() or "").strip()
        except Exception:
            subject_filter = ""

        selected_iid = None

        for idx_row, rec in enumerate(self._all_records):
            # filtr po roku szkolnym
            if year_filter and year_filter != "Wszystkie lata":
                if rec.get("school_year") != year_filter:
                    continue

            # filtr po przedmiocie – używamy klucza przedmiotu (subject_filter); pusty oznacza „wszystkie”
            if subject_filter:
                subj_rec = (rec.get("subject") or "").strip()
                if subj_rec != subject_filter:
                    continue

            combined = " ".join(
                [
                    rec.get("created", ""),
                    rec.get("context", ""),
                    rec.get("class_name", ""),
                    rec.get("subject", ""),
                    rec.get("school", ""),
                    rec.get("title", ""),
                    rec.get("students", ""),
                ]
            ).lower()

            if filter_text:
                if filter_text not in combined:
                    # tolerancja na odmianę (np. Kowalski / Kowalskie)
                    if len(filter_text) > 4 and filter_text[:-1] not in combined:
                        continue

            iid = f"I{idx_row}"
            self.tree_tests.insert(
                "",
                "end",
                iid=iid,
                values=(
                    rec["created"],
                    rec["context"],
                    rec["class_name"],
                    rec.get("subject", ""),
                    rec.get("school", ""),
                    rec["title"],
                ),
            )
            self._items_index[iid] = rec["path"]

            if self._preselect_path is not None and rec["path"] == self._preselect_path:
                selected_iid = iid

        if selected_iid is not None:
            self.tree_tests.selection_set(selected_iid)
            self.tree_tests.focus(selected_iid)
            self._on_select_item()
        else:
            children = self.tree_tests.get_children()
            if children:
                self.tree_tests.selection_set(children[0])
                self.tree_tests.focus(children[0])
                self._on_select_item()
            else:
                self._set_table(None, None)


    def _on_subject_tab_changed(self, event=None):
        """
        Reakcja na zmianę zakładki z przedmiotami.
        Ustawia filtr przedmiotu (subject_filter) na klucz przedmiotu powiązany z zakładką.
        Zakładka „Wszystkie” czyści filtr.
        """
        nb = getattr(self, "_subject_notebook", None)
        if nb is None:
            return
        try:
            tab_id = nb.select()
        except Exception:
            return
        if not tab_id:
            return
        try:
            text = nb.tab(tab_id, "text")
        except Exception:
            return
        label = (text or "").strip()
        # zakładka „Wszystkie” – brak filtra po przedmiocie
        if not label or label == "Wszystkie":
            try:
                self._subject_filter.set("")
            except Exception:
                return
            # przebuduj widok z nowym filtrem
            self._rebuild_treeview()
            return

        # odszukaj klucz przedmiotu po etykiecie w konfiguracji zakładek
        cfg = self._subject_tabs_cfg or {}
        tabs_cfg = cfg.get("tabs") or {}
        subj_key = ""
        for key, info in tabs_cfg.items():
            info_label = (info.get("label") or key).strip() if isinstance(info, dict) else str(info).strip()
            if info_label == label:
                subj_key = key
                break

        try:
            self._subject_filter.set(subj_key)
        except Exception:
            return

        # przebuduj widok z uwzględnieniem nowego filtra
        self._rebuild_treeview()

    def _show_student_overview_for_filter(self, text: str) -> bool:
        """
        Buduje po prawej stronie zbiorczy widok wszystkich sprawdzianów ucznia,
        którego nazwisko / imię i nazwisko wpisano w polu filtra.

        Zwraca:
            True  - jeśli znaleziono co najmniej jeden wiersz dla tego ucznia,
            False - jeśli nie znaleziono żadnego dopasowania (wtedy pozostaje widok standardowy).
        """
        pattern = (text or "").strip().lower()
        if not pattern:
            return False

        rows_out: list[dict] = []

        for rec in self._all_records:
            path = rec.get("path")
            if not path or not Path(path).exists():
                continue
            try:
                data = json.loads(Path(path).read_text(encoding="utf-8"))
            except Exception:
                continue

            cols = data.get("columns") or []
            rows = data.get("rows") or []
            meta = data.get("meta") or {}
            context_name = data.get("context", "") or ""
            class_name = ""
            if isinstance(meta, dict):
                class_name = str(meta.get("class_name", "") or "")
            title = data.get("title", Path(path).stem)

            # zbuduj DataFrame tak jak przy zwykłym podglądzie
            try:
                df = pd.DataFrame(rows, columns=cols) if cols else pd.DataFrame()
            except Exception:
                continue

            if df.empty:
                continue

            # ujednolicenie nazw kolumn (jak w sanitize_and_recompute)
            aliases = {
                "nazwisko": "Nazwisko",
                "imię i nazwisko": "Nazwisko",
                "imie i nazwisko": "Nazwisko",
                "ilosc punktow": "Ilość punktów",
                "ilość punktów": "Ilość punktów",
                "punkty": "Ilość punktów",
            }
            df.columns = [aliases.get(str(c).lower().strip(), str(c)) for c in df.columns]

            if "Nazwisko" not in df.columns:
                continue

            # filtrujemy tylko te wiersze, w których nazwisko zawiera wpisany wzorzec
            mask_student = df["Nazwisko"].astype(str).str.lower().str.contains(pattern)
            df_s = df[mask_student].copy()
            if df_s.empty:
                continue

            # kolumny pomocnicze – jeśli nie ma, dajemy puste
            col_points = "Ilość punktów" if "Ilość punktów" in df_s.columns else None
            col_percent = "Procent" if "Procent" in df_s.columns else None
            col_grade = "Ocena" if "Ocena" in df_s.columns else None

            for _, r in df_s.iterrows():
                rows_out.append(
                    {
                        "Uczeń": str(r.get("Nazwisko", "")),
                        "Data": rec.get("created", ""),
                        "Kontekst": rec.get("context", ""),
                        "Klasa / grupa": class_name,
                        "Tytuł sprawdzianu": title,
                        "Punkty": float(r.get(col_points)) if col_points else "",
                        "Procent": float(r.get(col_percent)) if col_percent else "",
                        "Ocena": r.get(col_grade, "") if col_grade else "",
                    }
                )

        if not rows_out:
            return False

        # budujemy zbiorczy DataFrame i pokazujemy go w tabeli po prawej
        df_overview = pd.DataFrame(rows_out)
        # nie potrzebujemy tu specjalnych meta – podajemy puste
        self._set_table(df_overview, meta={})
        return True

    def _on_select_item(self, event=None):
        sel = self.tree_tests.selection()
        if not sel:
            self._set_table(None, None)
            return
        iid = sel[0]
        path = self._items_index.get(iid)
        # zapamiętaj ostatnio wybrany plik jako preferowany przy filtrowaniu
        if path is not None:
            self._preselect_path = path
        if not path or not path.exists():
            self._set_table(None, None)
            return

        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            self._set_table(None, None)
            return

        cols = data.get("columns") or []
        rows = data.get("rows") or []
        meta = data.get("meta") or {}
        context_name = data.get("context", "")
        class_name = ""
        if isinstance(meta, dict):
            class_name = str(meta.get("class_name", "") or "")

        df = pd.DataFrame(rows, columns=cols) if cols else pd.DataFrame()
        if not df.empty:
            df.insert(0, "Kontekst (szkoła/placówka)", context_name)
            df.insert(1, "Klasa / grupa", class_name)
        self._set_table(df, meta)


    def _filter_by_selected_student(self):
        """
        Pobiera ucznia z zaznaczonego wiersza w tabeli wyników po prawej
        i ustawia:
        - filtr archiwum (po lewej) tak, aby pokazać wszystkie sprawdziany zawierające tego ucznia,
        - filtr ucznia po prawej (self._student_filter), aby w tabeli wyników pokazywać tylko jego wiersze.
        """
        sel = self.table.selection()
        if not sel:
            return
        iid = sel[0]
        values = self.table.item(iid, "values") or ()
        cols = list(self.table["columns"])

        # szukamy kolumny z nazwiskiem / imieniem i nazwiskiem
        idx_name = None
        for i, col in enumerate(cols):
            col_l = str(col).lower().strip()
            if col_l in ("nazwisko", "imię i nazwisko", "imie i nazwisko"):
                idx_name = i
                break
            if "nazwisk" in col_l or ("imi" in col_l and "nazw" in col_l) or "uczeń" in col_l or "uczen" in col_l:
                idx_name = i
                break

        if idx_name is None or idx_name >= len(values):
            return

        name_val = values[idx_name]
        if name_val is None:
            return
        name_text = str(name_val).strip()
        if not name_text:
            return

        # zapamiętujemy filtr ucznia (dla _set_table)
        self._student_filter = name_text

        # ustawiamy filtr tekstowy w archiwum – to wywoła _apply_filter/_rebuild_treeview
        # a następnie _on_select_item -> _set_table, które użyje self._student_filter
        self._filter_text.set(name_text)


    def _show_student_history(self):
        """
        Buduje przekrojowy raport „Historia ucznia” na podstawie całego archiwum.
        Uczeń wybierany jest z zaznaczonego wiersza lub – jeśli brak zaznaczenia – z okna dialogowego.
        """
        sel = self.table.selection()
        name_text = None
        if sel:
            iid = sel[0]
            values = self.table.item(iid, "values") or ()
            cols = list(self.table["columns"])
            idx_name = None
            for i, col in enumerate(cols):
                col_l = str(col).lower().strip()
                if col_l in ("nazwisko", "imię i nazwisko", "imie i nazwisko"):
                    idx_name = i
                    break
                if "nazwisk" in col_l or ("imi" in col_l and "nazw" in col_l) or "uczeń" in col_l or "uczen" in col_l:
                    idx_name = i
                    break
            if idx_name is not None and idx_name < len(values):
                val = values[idx_name]
                if val is not None:
                    sval = str(val).strip()
                    if sval:
                        name_text = sval

        if not name_text:
            from tkinter.simpledialog import askstring as _askstring
            name_text = _askstring(ARCHIVE_TITLE, "Podaj nazwisko (lub fragment) ucznia:")
            if not name_text:
                return

        arch_dir = _archive_dir()
        files = sorted(arch_dir.glob("*.json"), key=lambda p: p.stat().st_mtime)

        records: list[dict] = []

        for path in files:
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                continue

            created = data.get("created", "") or ""
            if "T" in created:
                created_disp = created.replace("T", " ")[:16]
            else:
                created_disp = created[:16]

            context = data.get("context", "") or ""
            title = data.get("title", "") or ""
            meta = data.get("meta") or {}
            class_name = ""
            if isinstance(meta, dict):
                class_name = str(meta.get("class_name", "") or "")

            cols_json = data.get("columns") or []
            rows_json = data.get("rows") or []
            if not cols_json or not rows_json:
                continue

            try:
                df = pd.DataFrame(rows_json, columns=[str(c) for c in cols_json])
            except Exception:
                continue

            col_name = None
            for col in df.columns:
                col_l = str(col).lower().strip()
                if col_l in ("nazwisko", "imię i nazwisko", "imie i nazwisko"):
                    col_name = col
                    break
                if "nazwisk" in col_l or ("imi" in col_l and "nazw" in col_l) or "uczeń" in col_l or "uczen" in col_l:
                    col_name = col
                    break
            if col_name is None:
                continue

            col_points = None
            col_percent = None
            col_grade = None
            for col in df.columns:
                cl = str(col).lower().strip()
                if col_points is None and ("punkt" in cl or "pkt" in cl):
                    col_points = col
                if col_percent is None and ("procent" in cl or "%" in cl):
                    col_percent = col
                if col_grade is None and "ocena" in cl:
                    col_grade = col

            try:
                mask = df[col_name].astype(str).str.contains(name_text, case=False, na=False)
                sub = df[mask].copy()
            except Exception:
                continue

            if sub.empty:
                continue

            for _, row in sub.iterrows():
                percent_display = ""
                if col_percent:
                    val = row.get(col_percent, "")
                    try:
                        if val not in ("", None):
                            v = float(str(val).replace(",", "."))
                            if 0.0 <= v <= 1.0001:
                                v *= 100.0
                            if float(v).is_integer():
                                percent_display = str(int(v))
                            else:
                                percent_display = f"{v:.2f}".replace(".", ",")
                    except Exception:
                        percent_display = str(val)

                rec = {
                    "Data": created_disp,
                    "Kontekst": context,
                    "Klasa": class_name,
                    "Przedmiot": title,
                    "Punkty": row.get(col_points, "") if col_points else "",
                    "Procent": percent_display,
                    "Ocena": row.get(col_grade, "") if col_grade else "",
                }
                records.append(rec)

        if not records:
            messagebox.showinfo(ARCHIVE_TITLE, f"Nie znaleziono wyników dla ucznia zawierającego: {name_text!r}.")
            return

        StudentHistoryWindow(self, name_text, records)

    def _set_table(self, df: pd.DataFrame | None, meta: dict | None):
        for col in self.table["columns"]:
            self.table.heading(col, text="")
            self.table.column(col, width=0)
        self.table["columns"] = ()
        for iid in self.table.get_children():
            self.table.delete(iid)

        # opcjonalne filtrowanie po konkretnym uczniu, jeśli ustawiono self._student_filter
        if df is not None and self._student_filter:
            try:
                col_name_for_filter = None
                for col in df.columns:
                    col_l = str(col).lower().strip()
                    if col_l in ("nazwisko", "imię i nazwisko", "imie i nazwisko"):
                        col_name_for_filter = col
                        break
                    if "nazwisk" in col_l or ("imi" in col_l and "nazw" in col_l) or "uczeń" in col_l or "uczen" in col_l:
                        col_name_for_filter = col
                        break
                if col_name_for_filter is not None:
                    mask = df[col_name_for_filter].astype(str).str.contains(self._student_filter, case=False, na=False)
                    df = df[mask].copy()
            except Exception:
                # w razie problemu z filtrowaniem nie zmieniamy df
                pass

        self._current_df = df
        self._current_meta = meta

        for item in self.scale_table.get_children():
            self.scale_table.delete(item)

        if df is None or df.empty:
            self.info_max_label.configure(text="Maksymalna liczba punktów (testu): –")
            self.info_method_label.configure(text="Metoda oceniania: –")
            self.info_weight_label.configure(text="Waga arkusza (jeśli użyto średniej ważonej): –")
            self.info_summary_label.configure(text="Podsumowanie wyników: –")
            return

        cols = [str(c) for c in df.columns]
        self.table["columns"] = cols

        # Najpierw wstaw dane, aby móc obliczyć szerokości na podstawie zawartości
        all_values = []
        for idx, row in df.iterrows():
            values = []
            for col in df.columns:
                val = row[col]
                if col == "Procent" and val is not None:
                    try:
                        pct = float(val) * 100.0
                        if float(pct).is_integer():
                            val = str(int(pct))
                        else:
                            val = f"{pct:.2f}".replace(".", ",")
                    except Exception:
                        pass
                values.append(str(val) if val is not None else "")
            all_values.append(values)
            grade_str = str(row.get("Ocena", "") or "")
            tags = ()
            if grade_str:
                d = grade_str.strip()[0]
                if d in self._grade_colors:
                    tags = (f"grade_{d}",)
            self.table.insert("", "end", iid=f"R{idx}", values=values, tags=tags)

        # Oblicz szerokości kolumn na podstawie zawartości (nazwa kolumny + najdłuższa wartość)
        for col_idx, col in enumerate(cols):
            self.table.heading(col, text=col)
            # Szerokość na podstawie nazwy kolumny
            header_width = len(col) * 8
            # Szerokość na podstawie najdłuższej wartości w kolumnie
            max_value_width = 0
            if all_values:
                for values in all_values:
                    if col_idx < len(values):
                        val_str = str(values[col_idx])
                        max_value_width = max(max_value_width, len(val_str) * 7)
            # Użyj większej z dwóch wartości, z minimalną szerokością 80 i maksymalną 300
            width = min(max(80, max(header_width, max_value_width + 10)), 300)
            self.table.column(col, width=width, anchor="w")

        max_points = None
        method_text = "–"
        scale_rows = None
        use_weighted = None
        sheet_weight = None

        if isinstance(meta, dict):
            max_points = meta.get("max_points")
            round_before = meta.get("round_before")
            scale_rows = meta.get("scale_rows")
            use_weighted = meta.get("use_weighted")
            sheet_weight = meta.get("sheet_weight")

            if round_before is True:
                method_text = "Metoda 1 – zaokrąglanie procentu"
            elif round_before is False:
                method_text = "Metoda 2 – bez zaokrąglania (lo ≤ % < hi+1)"
            else:
                method_text = "brak informacji"

        if max_points is not None:
            try:
                mp = float(max_points)
                if mp.is_integer():
                    max_text = f"Maksymalna liczba punktów (testu): {int(mp)}"
                else:
                    max_text = f"Maksymalna liczba punktów (testu): {mp}"
            except Exception:
                max_text = f"Maksymalna liczba punktów (testu): {max_points}"
        else:
            max_text = "Maksymalna liczba punktów (testu): brak informacji"

        self.info_max_label.configure(text=max_text)
        self.info_method_label.configure(text=f"Metoda oceniania: {method_text}")

        if use_weighted:
            if sheet_weight is not None:
                try:
                    w = float(sheet_weight)
                    if float(w).is_integer():
                        w_text = str(int(w))
                    else:
                        w_text = f"{w:.2f}".replace(".", ",")
                except Exception:
                    w_text = str(sheet_weight)
                weight_text = f"Waga arkusza (jeśli użyto średniej ważonej): {w_text}"
            else:
                weight_text = "Waga arkusza (jeśli użyto średniej ważonej): brak informacji"
        else:
            weight_text = "Waga arkusza (jeśli użyto średniej ważonej): nie dotyczy"

        self.info_weight_label.configure(text=weight_text)

        # podsumowanie wyników na podstawie meta["short_summary"]
        summary_text = "Podsumowanie wyników: –"
        if isinstance(meta, dict):
            short = meta.get("short_summary")
            if isinstance(short, str) and short.strip():
                summary_text = f"Podsumowanie wyników: {short}"
        self.info_summary_label.configure(text=summary_text)

        if isinstance(scale_rows, list) and scale_rows:
            for lo, hi, label in scale_rows:
                self.scale_table.insert("", "end", values=(lo, hi, label))
        else:
            self.scale_table.insert(
                "",
                "end",
                values=("-", "-", "Brak informacji o progach (starszy zapis)"),
            )

    def _get_selected_path(self) -> Path | None:
        sel = self.tree_tests.selection()
        if not sel:
            messagebox.showinfo(ARCHIVE_TITLE, "Najpierw zaznacz sprawdzian na liście.")
            return None
        iid = sel[0]
        path = self._items_index.get(iid)
        if not path or not path.exists():
            messagebox.showerror(ARCHIVE_TITLE, "Wybrany plik nie istnieje na dysku.")
            return None
        return path

    def _export_selected(self):
        path = self._get_selected_path()
        if not path:
            return

        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception as e:
            messagebox.showerror(ARCHIVE_TITLE, f"Nie mogę odczytać pliku:\n{path}\n\n{e}")
            return

        cols = data.get("columns") or []
        rows = data.get("rows") or []
        df = pd.DataFrame(rows, columns=cols) if cols else pd.DataFrame()

        if df.empty:
            messagebox.showinfo(ARCHIVE_TITLE, "Brak danych do eksportu.")
            return

        # Dodaj kolumny z kontekstem i klasą/grupą do eksportu
        context_name = data.get("context", "") or ""
        meta = data.get("meta") or {}
        class_name = ""
        if isinstance(meta, dict):
            class_name = str(meta.get("class_name", "") or "")
        df.insert(0, "Kontekst (szkoła/placówka)", context_name)
        df.insert(1, "Klasa / grupa", class_name)

        suggested_name = path.stem + ".xlsx"
        out_path = filedialog.asksaveasfilename(
            title="Zapisz wyniki jako…",
            defaultextension=".xlsx",
            initialfile=suggested_name,
            filetypes=[("Excel XLSX", "*.xlsx")],
        )
        if not out_path:
            return

        try:
            df.to_excel(out_path, index=False)
        except Exception as e:
            messagebox.showerror(ARCHIVE_TITLE, f"Nie udało się zapisać pliku:\n{e}")
            return

        messagebox.showinfo(ARCHIVE_TITLE, f"Zapisano wyniki do pliku:\n{out_path}")

    def _delete_selected(self):
        path = self._get_selected_path()
        if not path:
            return

        if not messagebox.askyesno(
            ARCHIVE_TITLE,
            f"Czy na pewno usunąć ten zapis z archiwum?\n\n{path.name}",
        ):
            return

        try:
            path.unlink(missing_ok=True)
        except Exception as e:
            messagebox.showerror(ARCHIVE_TITLE, f"Nie udało się usunąć pliku:\n{e}")
            return

        self._refresh_list()


def open_archive_window(master=None, preselect: Path | None = None) -> ArchiveViewer:
    return ArchiveViewer(master, preselect=preselect)


# ---------- profile skal (szablony) ----------
DEFAULT_SCALE = [
    (97, 100, "6 (celujący)"),
    (90, 96, "5 (bardzo dobry)"),
    (75, 89, "4 (dobry)"),
    (50, 74, "3 (dostateczny)"),
    (35, 49, "2 (dopuszczający)"),
    (0, 34, "1 (niedostateczny)"),
]

SCALE_PROFILES = {
    "Domyślna": DEFAULT_SCALE,
    "Sprawdzian 25 pkt": DEFAULT_SCALE,
    "Kartkówka 10 pkt": DEFAULT_SCALE,
}


# ---------- KONTEKSTY (szkoły) ----------
def _new_ctx_defaults() -> dict:
    return {
        "active_scale_rows": None,
        "scale_active": "Domyślna",
        "custom_scales": {},
        "max_points": 60,
        "open_after": True,
        "last_file": "",
        "last_output_dir": "",
        "use_weighted_mean": False,
        "weights_by_sheet": {},
        "weight_profiles": {},
        "active_weight_profile": "",
        "round_percent_before_grade": False,
    }


def _ensure_cfg_structure(cfg: dict) -> dict:
    if "contexts" not in cfg or not isinstance(cfg.get("contexts"), dict):
        cfg["contexts"] = {}
    if "current_context" not in cfg:
        cfg["current_context"] = "Domyślny"
    if cfg["current_context"] not in cfg["contexts"]:
        cfg["contexts"][cfg["current_context"]] = _new_ctx_defaults()
    for _n, ctx in cfg["contexts"].items():
        ctx.setdefault("use_weighted_mean", False)
        ctx.setdefault("weights_by_sheet", {})
        ctx.setdefault("weight_profiles", {})
        ctx.setdefault("active_weight_profile", "")
        ctx.setdefault("round_percent_before_grade", False)

    # globalne ustawienie motywu UI (jasny / ciemny)
    if "ui_theme" not in cfg:
        cfg["ui_theme"] = "light"  # domyślnie tryb jasny

    return cfg


def get_current_ctx_name(cfg: dict) -> str:
    cfg = _ensure_cfg_structure(cfg)
    return cfg["current_context"]


def get_ctx(cfg: dict, name: str | None = None) -> dict:
    cfg = _ensure_cfg_structure(cfg)
    ctx_name = name or cfg["current_context"]
    if ctx_name not in cfg["contexts"]:
        cfg["contexts"][ctx_name] = _new_ctx_defaults()
    return cfg["contexts"][ctx_name]


def set_ctx(cfg: dict, name: str, ctx: dict):
    cfg = _ensure_cfg_structure(cfg)
    ctx.setdefault("use_weighted_mean", False)
    ctx.setdefault("weights_by_sheet", {})
    ctx.setdefault("weight_profiles", {})
    ctx.setdefault("active_weight_profile", "")
    ctx.setdefault("round_percent_before_grade", False)
    cfg["contexts"][name] = ctx
    save_cfg(cfg)


def switch_ctx(cfg: dict, name: str) -> dict:
    cfg = _ensure_cfg_structure(cfg)
    if name not in cfg["contexts"]:
        cfg["contexts"][name] = _new_ctx_defaults()
    cfg["current_context"] = name
    save_cfg(cfg)
    return cfg


# ---------- skala aktywna ----------
def active_scale_from_ctx(cfg: dict) -> list[tuple]:
    ctx = get_ctx(cfg)
    if ctx.get("active_scale_rows"):
        return [(int(a), int(b), str(lbl)) for a, b, lbl in ctx["active_scale_rows"]]
    name = ctx.get("scale_active", "Domyślna")
    custom = (ctx.get("custom_scales") or {}).get(name)
    base = custom or SCALE_PROFILES.get(name, DEFAULT_SCALE)
    return [(int(a), int(b), str(lbl)) for a, b, lbl in base]


def set_active_scale_rows(cfg: dict, scale_rows: list[tuple], label: str):
    ctx = get_ctx(cfg).copy()
    ctx["active_scale_rows"] = [(int(a), int(b), str(lbl)) for a, b, lbl in scale_rows]
    ctx["scale_active"] = label
    set_ctx(cfg, get_current_ctx_name(cfg), ctx)


# ---------- logika ocen ----------
def compute_grade_from_percent(pct: float, scale_rows: list[tuple]) -> str:
    for lo, hi, label in scale_rows:
        if lo <= pct < hi + 1:
            return label
    return scale_rows[-1][2]


def grade_from_fraction(pct_frac: float, scale_rows: list[tuple], round_before: bool) -> str:
    pct = pct_frac * 100.0
    if round_before:
        pct = round(pct)
    return compute_grade_from_percent(float(pct), scale_rows)


# ---------- wczytywanie i przeliczanie ----------
def _detect_ext(path: str) -> str:
    return Path(path).suffix.lower()


def _excel_engine_for_ext(ext: str) -> dict:
    if ext == ".xlsx":
        return {"engine": "openpyxl"}
    if ext == ".xls":
        try:
            import xlrd  # noqa: F401
        except Exception:
            raise RuntimeError("Do plików .xls doinstaluj pakiet: pip install xlrd==1.2.0")
        return {"engine": "xlrd"}
    if ext == ".ods":
        try:
            import odf  # noqa: F401
        except Exception:
            raise RuntimeError("Do plików .ods doinstaluj pakiet: pip install odfpy")
        return {"engine": "odf"}
    return {}


def _normalize_loaded_df(df_in: pd.DataFrame) -> pd.DataFrame:
    try_df = df_in.copy()
    cols = [str(c).strip().lower() for c in try_df.columns]
    has_name = ("nazwisko" in cols) or ("imię i nazwisko" in cols) or ("imie i nazwisko" in cols)
    has_points = ("ilość punktów" in cols) or ("ilosc punktow" in cols) or ("punkty" in cols)
    if has_name and has_points:
        ren = {}
        for c in try_df.columns:
            cl = str(c).strip().lower()
            if cl in {"imię i nazwisko", "imie i nazwisko"}:
                ren[c] = "Nazwisko"
            if cl in {"ilość punktów", "ilosc punktow", "punkty"}:
                ren[c] = "Ilość punktów"
        if ren:
            try_df = try_df.rename(columns=ren)
        return try_df

    df2 = df_in.copy()
    df2.columns = [f"Kol{i}" for i in range(1, df2.shape[1] + 1)]
    cols2 = ["Nazwisko", "Ilość punktów"] + [f"Kol{i}" for i in range(3, df2.shape[1] + 1)]
    df2.columns = cols2[: df2.shape[1]]
    return df2


def _read_sheet_to_df(path: str, sheet_name):
    ext = _detect_ext(path)
    engine_kw = _excel_engine_for_ext(ext) if ext != ".csv" else {}
    
    # Najpierw wczytaj wszystkie wiersze jako dane (header=None)
    # aby nie stracić pierwszego wiersza, jeśli zawiera dane
    if ext == ".csv":
        raw_all = pd.read_csv(path, header=None)
    else:
        try:
            raw_all = pd.read_excel(path, sheet_name=sheet_name, header=None, **engine_kw)
        except Exception as e:
            # Fallback: spróbuj wczytać z domyślnym nagłówkiem
            try:
                raw_all = pd.read_excel(path, sheet_name=sheet_name, **engine_kw)
                return _normalize_loaded_df(raw_all)
            except Exception:
                raise e
    
    if raw_all.empty:
        return pd.DataFrame()
    
    # Sprawdź, czy pierwszy wiersz wygląda jak nagłówek (zawiera słowa kluczowe)
    first_row = raw_all.iloc[0]
    first_row_str = " ".join([str(x).strip().lower() for x in first_row.values if pd.notna(x)])
    header_keywords = ["nazwisko", "imię", "imie", "ilość punktów", "ilosc punktow", "punkty", "ocena", "procent"]
    looks_like_header = any(keyword in first_row_str for keyword in header_keywords)
    
    if looks_like_header:
        # Pierwszy wiersz to nagłówek - użyj go jako nagłówka
        raw_all.columns = [str(x).strip() for x in first_row.values]
        raw_all = raw_all.iloc[1:].reset_index(drop=True)
    else:
        # Pierwszy wiersz to dane - użyj domyślnych nazw kolumn
        raw_all.columns = [f"Kol{i}" for i in range(1, raw_all.shape[1] + 1)]
    
    return _normalize_loaded_df(raw_all)


def read_input_frames(path: str) -> dict:
    ext = _detect_ext(path)
    if ext == ".csv":
        name = Path(path).stem or "CSV"
        return {name: _read_sheet_to_df(path, None)}

    engine_kw = _excel_engine_for_ext(ext)
    xfile = pd.ExcelFile(path, **engine_kw)
    names = list(xfile.sheet_names)
    if not names:
        raise ValueError("W skoroszycie nie znaleziono żadnych arkuszy.")
    out = {}
    for s in names:
        out[s] = _read_sheet_to_df(path, s)
    return out


def sanitize_and_recompute(df: pd.DataFrame, max_points: float, scale_rows: list[tuple], round_before: bool) -> pd.DataFrame:
    df = df.copy()
    initial_count = len(df)
    df.columns = [str(c).strip() for c in df.columns]
    aliases = {
        "nazwisko": "Nazwisko",
        "imię i nazwisko": "Nazwisko",
        "imie i nazwisko": "Nazwisko",
        "ilosc punktow": "Ilość punktów",
        "ilość punktów": "Ilość punktów",
        "punkty": "Ilość punktów",
        "lp": "Lp.",
        "l.p.": "Lp.",
    }
    df = df.rename(columns={c: aliases.get(c.lower(), c) for c in df.columns})
    df = df.loc[:, ~df.columns.duplicated(keep="first")]

    if "Nazwisko" not in df.columns or "Ilość punktów" not in df.columns:
        raise ValueError("W pliku muszą być kolumny: 'Nazwisko' i 'Ilość punktów'.")

    df["Nazwisko"] = df["Nazwisko"].astype(str).str.strip()
    df["Ilość punktów"] = pd.to_numeric(df["Ilość punktów"], errors="coerce")
    
    # DIAGNOSTYKA: wyświetl co jest w NaN przed dropna
    nan_rows = df[df["Ilość punktów"].isna()]
    if not nan_rows.empty:
        print(f"[DEBUG] Wiersze z NaN w 'Ilość punktów' (będą usunięte): {list(nan_rows['Nazwisko'])}")
    
    # Usuń tylko wiersze z brakującymi punktami, ale zachowaj Nazwiska
    df = df.dropna(subset=["Ilość punktów"])
    df = df[df["Nazwisko"] != ""]
    
    after_cleanup = len(df)
    if initial_count != after_cleanup:
        print(f"[DEBUG] Wiersze: wejście {initial_count} → wyjście {after_cleanup} (usunięte: {initial_count - after_cleanup})")

    # --- Szybkie ostrzeżenia o podejrzanych punktach ---
    try:
        over_mask = df["Ilość punktów"] > max_points
        zero_mask = df["Ilość punktów"] == 0

        over_df = df.loc[over_mask, ["Nazwisko", "Ilość punktów"]]
        zero_df = df.loc[zero_mask, ["Nazwisko", "Ilość punktów"]]

        # 1) Uczniowie z punktami powyżej maksymalnej liczby
        if not over_df.empty:
            lines = [
                f"- {name} – {pts} pkt (max {max_points})"
                for name, pts in zip(over_df["Nazwisko"], over_df["Ilość punktów"])
            ]
            msg = (
                "Niektórzy uczniowie mają więcej punktów niż maksymalna liczba punktów.\n"
                "Sprawdź, czy nie ma literówek:\n\n" + "\n".join(lines)
            )
            try:
                from tkinter import messagebox  # lokalny import na wszelki wypadek
                messagebox.showerror(APP_TITLE, msg)
            except Exception:
                print(msg)

        # 2) Uczniowie z 0 punktów – łagodne ostrzeżenie
        if not zero_df.empty:
            lines = [
                f"- {name} – {pts} pkt"
                for name, pts in zip(zero_df["Nazwisko"], zero_df["Ilość punktów"])
            ]
            msg = (
                "Niektórzy uczniowie mają 0 punktów.\n"
                "Jeśli to nie jest celowe, sprawdź wprowadzone dane:\n\n"
                + "\n".join(lines)
            )
            try:
                from tkinter import messagebox  # lokalny import na wszelki wypadek
                messagebox.showwarning(APP_TITLE, msg)
            except Exception:
                print(msg)
    except Exception:
        # Ostrzeżenia nie mogą przerwać przeliczania – ignorujemy błędy w tej części
        pass

    # --- Liczenie procentów, ocen i porządkowanie danych ---
    df["Procent"] = df["Ilość punktów"] / max_points
    df["Ocena"] = df["Procent"].apply(
        lambda frac: grade_from_fraction(float(frac), scale_rows, round_before)
    )

    df = df.sort_values(by="Ilość punktów", ascending=False).reset_index(drop=True)
    if "Lp." in df.columns:
        df = df.drop(columns=["Lp."])
    df.insert(0, "Lp.", df.index + 1)

    wanted = ["Lp.", "Nazwisko", "Ilość punktów", "Procent", "Ocena"]
    others = [c for c in df.columns if c not in wanted]
    return df[wanted + others]



def _format_sheet(ws):
    widths = {"A": 6, "B": 22, "C": 16, "D": 12, "E": 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    # Nagłówki są w wierszu 1
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    # Zwiększ wysokość wiersza nagłówka, aby nie przykrywał pierwszego wiersza danych
    ws.row_dimensions[1].height = 20
    # Zamraź nagłówek (wiersz 1) — widz wiersze od 2 w dół
    ws.freeze_panes = ws["A2"]
    ws.auto_filter.ref = ws.dimensions
    for cell in ws["D"][1:]:
        cell.number_format = "0.00%"
    for cell in ws["E"]:
        cell.alignment = Alignment(horizontal="right")

    colors = {
        "6": "C6EFCE",
        "5": "CCFFCC",
        "4": "FFF2CC",
        "3": "FFD966",
        "2": "F4CCCC",
        "1": "EA9999",
    }
    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in row:
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        txt = row[4].value
        if isinstance(txt, str) and txt:
            d = txt[0]
            if d in colors:
                fill = PatternFill("solid", fgColor=colors[d])
                for c in row:
                    c.fill = fill


def _autofit_columns(ws, min_width=10, max_width=52):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            if v is None:
                continue
            s = str(v)
            max_len = max(max_len, int(len(s) * 1.1) + 2)
        width = max(min_width, min(max_width, max_len))
        ws.column_dimensions[letter].width = width


def _autofit_rows(ws, base_height=15, line_height=14):
    for r in range(1, ws.max_row + 1):
        lines = 1
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            lc = str(v).count("\n") + 1
            lines = max(lines, lc)
        ws.row_dimensions[r].height = max(base_height, lines * line_height)


def _append_scale_block(ws, scale_rows: list[tuple], round_before: bool, max_points: float):
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    start_row = ws.max_row + 2

    ws.cell(row=start_row, column=1, value="Informacje o ocenianiu:").font = Font(bold=True)
    ws.cell(row=start_row, column=1).fill = header_fill

    ws.cell(
        row=start_row + 1,
        column=1,
        value=f"Maksymalna liczba punktów (testu): {int(max_points) if float(max_points).is_integer() else max_points}",
    )
    ws.cell(
        row=start_row + 2,
        column=1,
        value="Metoda oceniania: "
        + (
            "Metoda 1 – zaokrąglanie procentu"
            if round_before
            else "Metoda 2 – bez zaokrąglania (lo ≤ % < hi+1)"
        ),
    )

    r = start_row + 4
    ws.cell(row=r - 1, column=1, value="Kryteria ocen (progi procentowe):").font = Font(bold=True)
    ws.cell(row=r - 1, column=1).fill = header_fill
    for lo, hi, label in scale_rows:
        ws.cell(row=r, column=1, value=f'({int(lo)}, {int(hi)}, "{label}")')
        r += 1


def _add_summary_sheet(
    wb,
    df: pd.DataFrame,
    after_sheet_name: str,
    title_suffix: str,
    scale_rows: list[tuple],
    round_before: bool,
    max_points: float,
):
    safe_after = after_sheet_name[:31]
    summary_title = f"Podsumowanie – {title_suffix}"[:31]
    try:
        pos = wb.sheetnames.index(safe_after)
    except ValueError:
        pos = len(wb.sheetnames) - 1
    s = wb.create_sheet(summary_title, index=pos + 1)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    row_fill = PatternFill("solid", fgColor="F2F2F2")
    row_border = Border(left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"), top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"))

    s["A1"], s["B1"] = "Ocena", "Liczba uczniów"
    s["D1"], s["E1"] = "Statystyka", "Wartość"
    for c in ["A1", "B1", "D1", "E1"]:
        s[c].font = Font(bold=True)
        s[c].fill = header_fill
        s[c].alignment = Alignment(horizontal="center")

    counts = df["Ocena"].astype(str).str[0].value_counts()
    r = 2
    for oc in ["6", "5", "4", "3", "2", "1"]:
        s[f"A{r}"] = oc
        s[f"B{r}"] = int(counts.get(oc, 0))
        r += 1

    # Pokoloruj wiersze tabeli rozkładu ocen
    for row_idx in range(2, r):
        for col in ("A", "B"):
            s[f"{col}{row_idx}"].fill = row_fill

    # Dodaj czarne ramki wokół komórek tabeli rozkładu ocen
    for row_idx in range(2, r):
        for col in ("A", "B"):
            s[f"{col}{row_idx}"].border = row_border

    stats = {
        "Średnia punktów": round(float(df["Ilość punktów"].mean()), 2) if len(df) else 0.0,
        "Mediana punktów": round(float(df["Ilość punktów"].median()), 2) if len(df) else 0.0,
        "Min punktów (uczeń)": int(df["Ilość punktów"].min()) if len(df) else 0,
        "Max punktów (uczeń)": int(df["Ilość punktów"].max()) if len(df) else 0,
        "Maksymalna liczba punktów (testu)": int(max_points)
        if float(max_points).is_integer()
        else float(max_points),
        "Liczba uczniów": int(len(df)),
        "Metoda oceniania": (
            "Metoda 1 – zaokrąglanie procentu"
            if round_before
            else "Metoda 2 – bez zaokrąglania (lo ≤ % < hi+1)"
        ),
    }
    r2 = 2
    for k, v in stats.items():
        s[f"D{r2}"] = k
        s[f"E{r2}"] = v
        r2 += 1

    # Pokoloruj wiersze tabeli statystyk
    for row_idx in range(2, r2):
        for col in ("D", "E"):
            s[f"{col}{row_idx}"].fill = row_fill

    # Dodaj czarne ramki wokół komórek tabeli statystyk
    for row_idx in range(2, r2):
        for col in ("D", "E"):
            s[f"{col}{row_idx}"].border = row_border

    for cell in s["E"][1:]:
        cell.number_format = "0.00" if isinstance(cell.value, float) else "0"
        cell.alignment = Alignment(horizontal="right")

    chart = BarChart()
    chart.title = f"Rozkład ocen – {title_suffix}"
    chart.y_axis.title = "Liczba uczniów"
    chart.x_axis.title = "Ocena"
    data_ref = Reference(s, min_col=2, min_row=1, max_row=7)
    cats_ref = Reference(s, min_col=1, min_row=2, max_row=7)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height, chart.width = 10, 18
    s.add_chart(chart, "A9")

    start_row = r2 + 2
    s[f"A{start_row}"] = "Kryteria ocen (progi procentowe):"
    s[f"A{start_row}"].font = Font(bold=True)
    s[f"A{start_row}"].fill = header_fill
    start_row += 1
    s[f"A{start_row}"], s[f"B{start_row}"], s[f"C{start_row}"] = "Od (%)", "Do (%)", "Ocena"
    for c in ["A", "B", "C"]:
        s[f"{c}{start_row}"].font = Font(bold=True)
        s[f"{c}{start_row}"].fill = header_fill
        s[f"{c}{start_row}"].alignment = Alignment(horizontal="center")

    row = start_row + 1
    for lo, hi, label in scale_rows:
        s[f"A{row}"] = lo
        s[f"B{row}"] = hi
        s[f"C{row}"] = label
        row += 1

    # Pokoloruj wiersze tabeli progów procentowych
    for row_idx in range(start_row + 1, row):
        for col in ("A", "B", "C"):
            s[f"{col}{row_idx}"].fill = row_fill

    # Dodaj czarne ramki wokół komórek progów procentowych
    for row_idx in range(start_row + 1, row):
        for col in ("A", "B", "C"):
            s[f"{col}{row_idx}"].border = row_border

    _autofit_columns(s)
    _autofit_rows(s)


def _weighted_mean_from_sheet_means(sheet_means: dict, weights: dict) -> float | None:
    num = 0.0
    den = 0.0
    for name, mean_val in sheet_means.items():
        w = float(weights.get(name, 1.0))
        if w <= 0:
            continue
        if pd.isna(mean_val):
            continue
        num += w * float(mean_val)
        den += w
    if den == 0:
        return None
    return num / den


def write_multi_with_formatting(
    sheet_dfs: dict[str, pd.DataFrame],
    out_path: str,
    use_weighted: bool,
    weights_by_sheet: dict,
    scale_rows: list[tuple],
    round_before: bool,
    max_points: float,
):
    order = list(sheet_dfs.keys())
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        for sname in order:
            # Zapisuj dane od wiersza 0 (nagłówki w 1, dane od 2)
            sheet_dfs[sname].to_excel(w, sheet_name=sname[:31], index=False, startrow=0)

    wb = load_workbook(out_path)
    sheet_means = {}

    for name in order:
        ws = wb[name[:31]]
        _format_sheet(ws)
        _append_scale_block(ws, scale_rows, round_before, max_points)
        _add_summary_sheet(wb, sheet_dfs[name], name, name, scale_rows, round_before, max_points)
        try:
            sheet_means[name] = float(sheet_dfs[name]["Ilość punktów"].mean())
        except Exception:
            sheet_means[name] = float("nan")

    all_df = pd.concat(sheet_dfs.values(), ignore_index=True) if sheet_dfs else pd.DataFrame()
    sname = "Zbiorcze podsumowanie"
    if sname in wb.sheetnames:
        del wb[sname]
    s = wb.create_sheet(sname)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    row_fill = PatternFill("solid", fgColor="F2F2F2")
    row_border = Border(left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"), top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"))

    s["A1"], s["B1"] = "Ocena", "Łącznie uczniów"
    s["D1"], s["E1"] = "Statystyka (globalnie)", "Wartość"
    for c in ["A1", "B1", "D1", "E1"]:
        s[c].font = Font(bold=True)
        s[c].fill = header_fill
        s[c].alignment = Alignment(horizontal="center")

    if not all_df.empty:
        counts = all_df["Ocena"].astype(str).str[0].value_counts()
        r = 2
        for oc in ["6", "5", "4", "3", "2", "1"]:
            s[f"A{r}"] = oc
            s[f"B{r}"] = int(counts.get(oc, 0))
            r += 1

        # Pokoloruj wiersze tabeli rozkładu ocen (zbiorcze)
        for row_idx in range(2, r):
            for col in ("A", "B"):
                s[f"{col}{row_idx}"].fill = row_fill

        # Dodaj czarne ramki wokół komórek tabeli rozkładu ocen (zbiorcze)
        for row_idx in range(2, r):
            for col in ("A", "B"):
                s[f"{col}{row_idx}"].border = row_border

        stats = {
            "Średnia punktów": round(float(all_df["Ilość punktów"].mean()), 2),
            "Mediana punktów": round(float(all_df["Ilość punktów"].median()), 2),
            "Min punktów (uczeń)": int(all_df["Ilość punktów"].min()),
            "Max punktów (uczeń)": int(all_df["Ilość punktów"].max()),
            "Maksymalna liczba punktów (testu)": int(max_points)
            if float(max_points).is_integer()
            else float(max_points),
            "Łącznie uczniów": int(len(all_df)),
            "Metoda oceniania": (
                "Metoda 1 – zaokrąglanie procentu"
                if round_before
                else "Metoda 2 – bez zaokrąglania (lo ≤ % < hi+1)"
            ),
        }
        if use_weighted:
            wm = _weighted_mean_from_sheet_means(sheet_means, weights_by_sheet or {})
            if wm is not None:
                stats["Średnia punktów (ważona)"] = round(float(wm), 2)

        r2 = 2
        for k, v in stats.items():
            s[f"D{r2}"] = k
            s[f"E{r2}"] = v
            r2 += 1

        # Pokoloruj wiersze tabeli statystyk (zbiorcze)
        for row_idx in range(2, r2):
            for col in ("D", "E"):
                s[f"{col}{row_idx}"].fill = row_fill

        # Dodaj czarne ramki wokół komórek tabeli statystyk (zbiorcze)
        for row_idx in range(2, r2):
            for col in ("D", "E"):
                s[f"{col}{row_idx}"].border = row_border

        for cell in s["E"][1:]:
            cell.number_format = "0.00" if isinstance(cell.value, float) else "0"
            cell.alignment = Alignment(horizontal="right")

        start_row = r2 + 2
        s[f"A{start_row}"] = "Kryteria ocen (progi procentowe):"
        s[f"A{start_row}"].font = Font(bold=True)
        s[f"A{start_row}"].fill = header_fill
        start_row += 1
        s[f"A{start_row}"], s[f"B{start_row}"], s[f"C{start_row}"] = "Od (%)", "Do (%)", "Ocena"
        for c in ["A", "B", "C"]:
            s[f"{c}{start_row}"].font = Font(bold=True)
            s[f"{c}{start_row}"].fill = header_fill
            s[f"{c}{start_row}"].alignment = Alignment(horizontal="center")
        row = start_row + 1
        for lo, hi, label in scale_rows:
            s[f"A{row}"] = lo
            s[f"B{row}"] = hi
            s[f"C{row}"] = label
            row += 1

        # Pokoloruj wiersze tabeli progów procentowych (zbiorcze)
        for row_idx in range(start_row + 1, row):
            for col in ("A", "B", "C"):
                s[f"{col}{row_idx}"].fill = row_fill

        # Dodaj czarne ramki wokół komórek progów procentowych (zbiorcze)
        for row_idx in range(start_row + 1, row):
            for col in ("A", "B", "C"):
                s[f"{col}{row_idx}"].border = row_border

    _autofit_columns(s)
    _autofit_rows(s)

    wb.save(out_path)


def process_file_all_sheets(
    in_path: str,
    max_points: float,
    out_path: str,
    scale_rows: list[tuple],
    use_weighted: bool,
    weights_by_sheet: dict,
    round_before: bool,
) -> dict[str, pd.DataFrame]:
    """
    Przetwarza wszystkie arkusze z pliku z wyjątkiem technicznych,
    takich jak META / _meta (służących np. do opisu: przedmiot, klasa, szkoła).
    """
    sheets_in = read_input_frames(in_path)
    result: dict[str, pd.DataFrame] = {}
    for sname, df_in in sheets_in.items():
        # Pomijamy arkusze techniczne META
        sname_norm = str(sname).strip().lower()
        if sname_norm in {"meta", "_meta"}:
            continue
        df_out = sanitize_and_recompute(df_in, max_points, scale_rows, round_before)
        result[sname] = df_out
    write_multi_with_formatting(result, out_path, use_weighted, weights_by_sheet, scale_rows, round_before, max_points)
    return result


# =============== GUI ===============
class App(ttk.Frame):
    BG = "#F4F6FB"
    FG = "#1F2937"
    ACCENT = "#1E88E5"
    ACCENT_H = "#1565C0"
    BORDER = "#E5E7EB"

    def __init__(self, root: TkBase):
        super().__init__(root)
        root.title(APP_TITLE)
        root.minsize(880, 680)

        # konfiguracja
        self.cfg = _ensure_cfg_structure(load_cfg())

        # Motyw interfejsu (jasny / ciemny) – domyślnie „light”
        self.ui_theme = self.cfg.get("ui_theme", "light")

        # zastosuj motyw
        self._apply_theme(root)

        # Domyślne wartości z konfiguracji (okno „Ustawienia programu”)
        default_school = self.cfg.get("default_school", "")
        default_subject = self.cfg.get("default_subject", "")
        default_class = self.cfg.get("default_class", "")

        self.ctx_name = tk.StringVar(value=get_current_ctx_name(self.cfg))
        # Klasa / grupa – opcjonalne pole tekstowe
        self.class_name = tk.StringVar(value=default_class)
        # Przedmiot – dowolny tekst (np. Historia, Matematyka)
        self.subject_var = tk.StringVar(value=default_subject)
        # Szkoła – opcjonalne pole (np. pełna nazwa placówki)
        self.school_var = tk.StringVar(value=default_school)

        ctx = get_ctx(self.cfg)
        self.file_path = tk.StringVar(value=ctx.get("last_file", ""))
        self.max_points = tk.StringVar(value=str(ctx.get("max_points", 60)))
        self.open_after = tk.BooleanVar(value=bool(ctx.get("open_after", True)))
        self.output_dir = tk.StringVar(value=ctx.get("last_output_dir", ""))
        self.use_weighted = tk.BooleanVar(value=bool(ctx.get("use_weighted_mean", False)))
        self.round_before = tk.BooleanVar(value=bool(ctx.get("round_percent_before_grade", False)))

        default_profile = ctx.get("scale_active", "Domyślna")
        self.active_scale_name = tk.StringVar(value=default_profile)

        self.batch_mode = tk.BooleanVar(value=False)
        self.batch_files = []

        self._last_archive_path: Path | None = None

        pad = 10

        # helper: tworzy zewnętrzną ramkę w kolorze oraz wewnętrzny LabelFrame
        # oraz dodaje subtelny 3D cień (offsetowany i ciemniejszy) pod ramką.
        def create_colored_section(text: str, color: str, pack_opts: dict | None = None, parent=None, **lf_kwargs):
            """
            parent: miejsce w którym zostanie wstawiona ramka (domyślnie root)
            pack_opts: dodatkowe opcje pack dla zewnętrznej ramki
            Zwraca wewnętrzny `ttk.LabelFrame` i przechowuje zewnętrzny wrapper
            w atrybucie `lf._outer` (tak jak wcześniej).
            """
            p = parent if parent is not None else root

            # Wrapper zawsze ma jasne tło (domyślny kolor aplikacji light mode)
            # aby uniknąć ciemnych odstępów między sekcjami
            wrapper_bg = "#F4F6FB"

            # prosty helper do ściemniania koloru (hex -> ciemniejszy hex)
            def _darken_hex(hx: str, factor: float = 0.55) -> str:
                try:
                    hx = hx.lstrip("#")
                    r = int(hx[0:2], 16)
                    g = int(hx[2:4], 16)
                    b = int(hx[4:6], 16)
                    r = max(0, int(r * factor))
                    g = max(0, int(g * factor))
                    b = max(0, int(b * factor))
                    return f"#{r:02x}{g:02x}{b:02x}"
                except Exception:
                    return "#000000"

            shadow_color = _darken_hex(color, 0.5)

            # wrapper: ten widget caller będzie pack/gridował (zamiast wcześniejszego outer)
            wrapper = tk.Frame(p, bg=wrapper_bg)

            # Jeśli parent to ttk.Frame lub zwykły tk.Frame (ale nie root), 
            # znaczy że to container na grid — nie rób pack
            is_grid_container = isinstance(p, ttk.Frame) or (isinstance(p, tk.Frame) and p != root)
            if not is_grid_container:
                opts = dict(fill="both", expand=True, padx=pad, pady=(0, pad))
                if pack_opts:
                    opts.update(pack_opts)
                wrapper.pack(**opts)

            # shadow umieszczamy jako pierwszy (będzie pod spodem).
            # Jeśli w folderze ze skryptem jest obraz `ab.png`, użyjemy go
            # jako cienia (ładniejszy efekt). W przeciwnym razie tworzymy prosty
            # ciemniejszy Frame jako cień.
            shadow_img_label = None
            try:
                script_dir = Path(__file__).parent
                img_path = script_dir / "ab.png"
                if img_path.exists():
                    try:
                        from PIL import Image, ImageTk

                        pil = Image.open(img_path)
                        shadow_img = ImageTk.PhotoImage(pil)
                        shadow_img_label = tk.Label(wrapper, image=shadow_img, bd=0)
                        # zachowaj referencję, aby GC nie usuwał obrazka
                        wrapper._shadow_img = shadow_img
                        try:
                            shadow_img_label.place(relx=0, rely=0, x=6, y=6, anchor="nw")
                        except Exception:
                            shadow_img_label.pack(fill="both", expand=True, padx=(6, 0), pady=(6, 0))
                    except Exception:
                        shadow_img_label = None
                else:
                    shadow_img_label = None
            except Exception:
                shadow_img_label = None

            if shadow_img_label is None:
                # cień: mały Frame tylko na dole-prawie (bez zajmowania całej przestrzeni)
                shadow = tk.Frame(wrapper, bg=shadow_color, bd=0, highlightthickness=0, width=6, height=6)
                # umieść cień dokładnie na dole-prawie za outer
                shadow.place(relx=1.0, rely=1.0, x=-6, y=-6, anchor="se")

            # kolorowa ramka (outer) zajmuje cały wrapper (bez padding)
            outer = tk.Frame(wrapper, bg=color, highlightthickness=3, highlightbackground=color, bd=0)
            outer.pack(fill="both", expand=True)

            # wewnętrzny LabelFrame jak wcześniej
            lf = ttk.LabelFrame(outer, text=text, padding=pad // 2, style="Flat.TLabelframe", **lf_kwargs)
            lf.pack(fill="both", expand=True, padx=6, pady=6)

            # Przechowujemy wrapper (nie bezpośrednio outer) — caller nadal używa lf._outer
            lf._outer = wrapper
            return lf

        top = ttk.Frame(root, padding=pad, style="Flat.TFrame")
        top.pack(fill="x")
        self._logo_img = None
        if Path(PATH_LOGO).exists():
            try:
                from PIL import Image, ImageTk

                img = Image.open(PATH_LOGO)
                img.thumbnail((110, 110))
                self._logo_img = ImageTk.PhotoImage(img)
                ttk.Label(top, image=self._logo_img, style="Flat.TLabel").pack(side="left", padx=(0, pad))
            except Exception:
                pass
        ttk.Label(
            top,
            text="Zespół Szkół w Górznie\nProgram „Wyniki 5”",
            font=("Segoe UI", 14, "bold"),
            style="Flat.TLabel",
            justify="left",
        ).pack(side="left", anchor="w")

        # Sekcja Kontekst – pełna szerokość na górze
        ctx_outer = ttk.Frame(root)
        ctx_outer.pack(fill="x", padx=pad, pady=(0, pad//2))
        lf_ctx = create_colored_section("Kontekst (szkoła/placówka)", "#7A9CFF", parent=ctx_outer)
        ttk.Label(lf_ctx, text="Bieżący kontekst:", style="Flat.TLabel").grid(row=0, column=0, sticky="w")
        self.cb_ctx = ttk.Combobox(
        lf_ctx,
        state="readonly",
        width=32,
        values=list(self.cfg["contexts"].keys()),
        textvariable=self.ctx_name,
        )
        self.cb_ctx.grid(row=0, column=1, sticky="w", padx=(6, 10))
        ttk.Button(lf_ctx, text="Nowy…", command=self._ctx_new, style="TButton").grid(row=0, column=2, sticky="w")
        ttk.Button(
        lf_ctx, text="Zmień nazwę / Usuń…", command=self._ctx_manage, style="TButton"
        ).grid(row=0, column=3, sticky="w", padx=(6, 0))
        ttk.Label(lf_ctx, text="Klasa / grupa:", style="Flat.TLabel").grid(row=1, column=0, sticky="w", pady=(6, 0))
        self.entry_class = ttk.Entry(lf_ctx, textvariable=self.class_name, width=15, style="Flat.TEntry")
        self.entry_class.grid(row=1, column=1, sticky="w", padx=(6, 10), pady=(6, 0))

        ttk.Label(lf_ctx, text="Przedmiot:", style="Flat.TLabel").grid(row=2, column=0, sticky="w", pady=(4, 0))
        self.entry_subject = ttk.Entry(lf_ctx, textvariable=self.subject_var, width=20, style="Flat.TEntry")
        self.entry_subject.grid(row=2, column=1, sticky="w", padx=(6, 10), pady=(4, 0))

        ttk.Label(lf_ctx, text="Szkoła (opcjonalnie):", style="Flat.TLabel").grid(row=3, column=0, sticky="w", pady=(4, 0))
        self.entry_school = ttk.Entry(lf_ctx, textvariable=self.school_var, width=32, style="Flat.TEntry")
        self.entry_school.grid(row=3, column=1, sticky="w", padx=(6, 10), pady=(4, 0))

        # Główny kontener siatka 2 kolumny × 3 wiersze
        main_grid = tk.Frame(root, bg="#F4F6FB")
        self.main_grid = main_grid
        main_grid.pack(fill="both", expand=True, padx=pad, pady=(0, pad))
        
        # Konfiguruj kolumny i wiersze aby były równomiernie rozszerzone
        main_grid.columnconfigure(0, weight=1)
        main_grid.columnconfigure(1, weight=1)
        main_grid.rowconfigure(0, weight=1)
        main_grid.rowconfigure(1, weight=1)
        main_grid.rowconfigure(2, weight=1)

        # Spacers on the right column to preserve row heights when widgets are hidden
        self._spacer_r1 = ttk.Frame(main_grid, style="Flat.TFrame")
        self._spacer_r2 = ttk.Frame(main_grid, style="Flat.TFrame")
        # Don't grid them now; _toggle_batch will show appropriate spacer when needed

        lf_batch = create_colored_section("Tryb działania", "#7AE7FF", parent=main_grid)
        # Ustaw grid dla outer frame (parent zawiera outer)
        # Wiersz 0, Kolumna 1
        lf_batch._outer.grid(row=0, column=1, sticky="nsew", padx=(pad//2, 0), pady=(0, 3))
        ttk.Checkbutton(
            lf_batch,
            text="Tryb wsadowy (wiele plików → jeden folder „wyniki”)",
            variable=self.batch_mode,
            command=self._toggle_batch,
            style="Flat.TCheckbutton",
        ).grid(row=0, column=0, sticky="w")

        self.lf_file = create_colored_section("Plik z wynikami (.xlsx/.xls/.ods/.csv)", "#FFDA7A", parent=main_grid)
        self.lf_file._outer.grid(row=1, column=1, sticky="nsew", padx=(pad//2, 0), pady=(3, 3))
        row1 = ttk.Frame(self.lf_file, style="Flat.TFrame")
        row1.pack(fill="x", pady=(0, 3))
        self.entry_file = ttk.Entry(row1, textvariable=self.file_path, style="Flat.TEntry")
        self.entry_file.pack(side="left", fill="x", expand=True)
        ttk.Button(row1, text="Wybierz…", command=self.pick_file, style="TButton").pack(side="left", padx=(pad//2, 0))
        info = "Możesz też przeciągnąć tu plik .xlsx, .xls, .ods lub .csv."
        if not USE_DND:
            info += " (zainstaluj tkinterdnd2, aby włączyć przeciąganie)"
        ttk.Label(self.lf_file, text=info, style="Flat.TLabel").pack(anchor="w", pady=(3, 2))
        if USE_DND:
            try:
                self.entry_file.drop_target_register(DND_FILES)
                self.entry_file.dnd_bind("<<Drop>>", self._on_drop_single)
            except Exception:
                pass

        self.lf_files = create_colored_section("Pliki wsadowe (.xlsx/.xls/.ods)", "#B8FF7A", parent=main_grid)
        self.lf_files._outer.grid(row=2, column=1, sticky="nsew", padx=(pad//2, 0), pady=(3, 0))
        self.lf_files._outer.grid_remove()  # Ukryj na starcie - pojawi się w trybie wsadowym
        rowb = ttk.Frame(self.lf_files, style="Flat.TFrame")
        self.btn_pick_files = ttk.Button(rowb, text="Wybierz pliki…", command=self.pick_files, style="TButton")
        self.btn_pick_files.pack(side="left")
        self.lbl_files = ttk.Label(rowb, text="(nie wybrano plików)", style="Flat.TLabel")
        self.lbl_files.pack(side="left", padx=(4, 0))
        rowb.pack(fill="x")

        # Folder wyjściowy umieszczony w siatce pod Pliki wsadowe (row=2, col=1)
        self.lf_out = create_colored_section("Folder wyjściowy", "#B8FF7A", parent=main_grid)
        # Umieść outer w grid (domyślnie ukryty)
        self.lf_out._outer.grid(row=2, column=1, sticky="nsew", padx=(pad//2, 0), pady=(3, 0))
        # Ukryj na starcie — pojawi się dopiero w trybie wsadowym
        self.lf_out._outer.grid_remove()
        rowo = ttk.Frame(self.lf_out, style="Flat.TFrame")
        self.entry_out = ttk.Entry(rowo, textvariable=self.output_dir, style="Flat.TEntry")
        self.entry_out.pack(side="left", fill="x", expand=True)
        ttk.Button(rowo, text="Wybierz folder…", command=self.pick_output_dir, style="TButton").pack(
            side="left", padx=(pad//2, 0)
        )
        rowo.pack(fill="x")

        lf_params = create_colored_section("Parametry", "#C27AFF", parent=main_grid)
        lf_params._outer.grid(row=0, column=0, sticky="nsew", padx=(0, pad//2), pady=(0, 3))
        ttk.Label(lf_params, text="Maksymalna liczba punktów (testu):", style="Flat.TLabel").grid(
            row=0, column=0, sticky="w"
        )
        ttk.Spinbox(
            lf_params, from_=1, to=1000, textvariable=self.max_points, width=10, style="Flat.TSpinbox"
        ).grid(row=0, column=1, sticky="w", padx=(3, 0))
        ttk.Checkbutton(
            lf_params,
            text="Otwórz plik po zapisaniu (tryb pojedynczy)",
            variable=self.open_after,
            style="Flat.TCheckbutton",
        ).grid(row=0, column=2, sticky="w", padx=(10, 0))
        ttk.Checkbutton(
            lf_params,
            text="Użyj średniej ważonej w zbiorczym podsumowaniu",
            variable=self.use_weighted,
            style="Flat.TCheckbutton",
        ).grid(row=1, column=0, columnspan=2, sticky="w", pady=(3, 0))
        ttk.Button(lf_params, text="Wagi arkuszy…", command=self.edit_weights, style="TButton").grid(
            row=1, column=2, sticky="e"
        )
        ttk.Button(lf_params, text="Profile wag…", command=self.edit_weight_profiles, style="TButton").grid(
            row=2, column=2, sticky="e", pady=(4, 0)
        )

        lf_method = create_colored_section("Metoda oceniania", "#FF7A7A", parent=main_grid)
        lf_method._outer.grid(row=1, column=0, sticky="nsew", padx=(0, pad//2), pady=(3, 3))
        ttk.Radiobutton(
            lf_method,
            text="Metoda 1 – zaokrąglaj procent przed oceną",
            value=True,
            variable=self.round_before,
            style="Flat.TRadiobutton",
        ).grid(row=0, column=0, sticky="w")
        ttk.Radiobutton(
            lf_method,
            text="Metoda 2 – bez zaokrąglania (lo ≤ % < hi+1)",
            value=False,
            variable=self.round_before,
            style="Flat.TRadiobutton",
        ).grid(row=1, column=0, sticky="w")

        # UWAGA: lf_scale musi być w rzędzie 2, ale left_col zniknął
        # Tymczasowo zostawiamy poniżej - będzie to ostatnia sekcja
        lf_scale = create_colored_section("Skala ocen (profile w bieżącym kontekście)", "#7AEEB1", parent=main_grid)
        lf_scale._outer.grid(row=2, column=0, sticky="nsew", padx=(0, pad//2), pady=(3, 0))
        ttk.Label(lf_scale, text="Profil:", style="Flat.TLabel").grid(row=0, column=0, sticky="w")
        profiles = set(SCALE_PROFILES.keys())
        profiles |= set((get_ctx(self.cfg).get("custom_scales") or {}).keys())
        profiles = list(sorted(profiles))
        active_label = self.active_scale_name.get()
        if active_label not in profiles:
            profiles.insert(0, active_label)
        self.cb_profile = ttk.Combobox(
            lf_scale, state="readonly", values=profiles, textvariable=self.active_scale_name, width=28
        )
        self.cb_profile.grid(row=0, column=1, sticky="w", padx=(6, 10))
        self.cb_profile.bind("<<ComboboxSelected>>", self._on_profile_selected)
        ttk.Button(lf_scale, text="Skala…", command=self.open_scale_editor, style="TButton").grid(
            row=0, column=2, sticky="w"
        )

        self.progress = ttk.Progressbar(root, mode="determinate", maximum=100)
        self.progress.pack(fill="x", padx=pad, pady=(0, pad//2))

        actions = ttk.Frame(root, padding=(pad//2, pad//2, pad//2, pad//2), style="Flat.TFrame")
        actions.pack(fill="x")
        self.btn_run = ttk.Button(actions, text="Przelicz i zapisz", command=self.run, style="Accent.TButton")
        self.btn_run.pack(side="left")
        ttk.Button(actions, text="Wpisz dane ręcznie…", command=self.open_manual_input, style="TButton").pack(
            side="left", padx=(pad//3, 0)
        )
        ttk.Button(actions, text="Archiwum wyników…", command=self.open_archive, style="TButton").pack(
            side="left", padx=(pad//3, 0)
        )
        ttk.Button(actions, text="Pokaż ostatni wynik", command=self.show_last_result, style="TButton").pack(
            side="left", padx=(pad//3, 0)
        )
        ttk.Button(actions, text="Ustawienia programu…", command=self.open_settings_window, style="TButton").pack(
            side="left", padx=(pad//3, 0)
        )
        ttk.Button(actions, text="Otwórz instrukcję", command=self.open_manual, style="TButton").pack(
            side="left", padx=(pad//3, 0)
        )

        self.status = tk.StringVar(value="Gotowy.")
        ttk.Label(root, textvariable=self.status, style="Flat.TLabel").pack(anchor="w", padx=pad, pady=(pad//2, 0))
        ttk.Label(
            root,
            text="© 2025 Zespół Szkół w Górznie · opracował: Jarek – SP Górzno",
            style="Flat.TLabel",
        ).pack(side="bottom", pady=(0, pad))

        self._toggle_batch(init=True)
        # Lock row heights so toggling batch widgets doesn't shift layout
        try:
            self._lock_main_grid_row_heights()
        except Exception:
            pass
        self.cb_ctx.bind("<<ComboboxSelected>>", lambda _e=None: self._reload_from_context())

        try:
            root.bind_all("<Control-o>", lambda e: self.pick_file())
            root.bind_all("<Control-s>", lambda e: self.run())
            root.bind_all("<Alt-r>", lambda e: self.run())
            root.bind_all("<F1>", lambda e: self.open_manual())
        except Exception:
            pass

    def _apply_theme(self, root: TkBase):
        """
        Bardzo nowoczesny wygląd z wyborem trybu:
        - self.ui_theme == "light" -> nowoczesny jasny,
        - self.ui_theme == "dark"  -> nowoczesny ciemny.
        Funkcjonalność programu pozostaje bez zmian.
        """

        # Bezpieczna wartość domyślna
        theme_mode = getattr(self, "ui_theme", "light")
        if theme_mode not in ("light", "dark"):
            theme_mode = "light"

        # ===== wariant z ttkbootstrap =====
        if 'USE_TTKB' in globals() and USE_TTKB and TtkbStyle is not None:
            try:
                if theme_mode == "light":
                    theme_name = "flatly"
                else:
                    theme_name = "superhero"

                style = TtkbStyle(theme=theme_name)

                # kolor tła okna głównego
                root.configure(background=style.master.cget("background"))

                # podstawowe style – zostawiamy nazwy jak w programie
                style.configure("Flat.TFrame", padding=4)
                style.configure("Flat.TLabelframe", padding=8)
                style.configure("Flat.TLabel", font=("Segoe UI", 10))
                style.configure("Title.TLabel", font=("Segoe UI", 14, "bold"))
                style.configure("Flat.TEntry", font=("Segoe UI", 10))
                style.configure("Flat.TSpinbox", font=("Segoe UI", 10))
                style.configure("Flat.TCheckbutton", font=("Segoe UI", 10))
                style.configure("Flat.TRadiobutton", font=("Segoe UI", 10))
                style.configure("TButton", font=("Segoe UI", 10, "bold"))
                style.configure("Status.TLabel", font=("Segoe UI", 9))
            except Exception:
                # jeśli coś pójdzie nie tak z ttkbootstrap – spadamy do zwykłego ttk
                pass

        # ===== wariant bez ttkbootstrap – ręczne style ttk =====
        if theme_mode == "light":
            bg = "#F4F6FB"
            fg = "#1F2937"
            accent = "#1E88E5"
            accent_h = "#1565C0"
            border = "#E5E7EB"
        else:
            # nowoczesny ciemny
            bg = "#111827"        # tło okna
            fg = "#F9FAFB"        # tekst
            accent = "#3B82F6"    # niebieski akcent
            accent_h = "#1D4ED8"
            border = "#374151"

        self.BG = bg
        self.FG = fg
        self.ACCENT = accent
        self.ACCENT_H = accent_h
        self.BORDER = border

        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass

        root.configure(bg=bg)

        style.configure("Flat.TFrame", background=bg)
        style.configure(
            "Flat.TLabelframe",
            background=bg,
            foreground=fg,
            bordercolor=border,
            relief="solid",
        )
        # Make the LabelFrame title use the same background and be bold (no darker title box)
        try:
            style.configure("Flat.TLabelframe.Label", background=bg, foreground=fg, font=("Segoe UI", 10, "bold"))
        except Exception:
            pass
        style.configure(
            "Flat.TLabel",
            background=bg,
            foreground=fg,
            font=("Segoe UI", 10),
        )
        style.configure(
            "Title.TLabel",
            background=bg,
            foreground=fg,
            font=("Segoe UI", 14, "bold"),
        )
        style.configure(
            "Flat.TEntry",
            fieldbackground="#FFFFFF" if theme_mode == "light" else "#1F2933",
            foreground=fg,
            bordercolor=border,
            relief="flat",
        )
        style.configure(
            "Flat.TSpinbox",
            fieldbackground="#FFFFFF" if theme_mode == "light" else "#1F2933",
            foreground=fg,
            bordercolor=border,
            relief="flat",
        )
        style.configure(
            "Flat.TCheckbutton",
            background=bg,
            foreground=fg,
            font=("Segoe UI", 10),
        )
        style.configure(
            "Flat.TRadiobutton",
            background=bg,
            foreground=fg,
            font=("Segoe UI", 10),
        )
        style.configure(
            "TButton",
            background=accent,
            foreground="#FFFFFF",
            focusthickness=1,
            focuscolor=accent_h,
            padding=6,
            font=("Segoe UI", 10, "bold"),
        )
        style.map(
            "TButton",
            background=[("active", accent_h)],
        )
        style.configure(
            "Status.TLabel",
            background=bg,
            foreground=fg,
            font=("Segoe UI", 9),
        )

    def _lock_main_grid_row_heights(self):
        """
        Measure current heights of widgets in each grid row and set
        `minsize` for `self.main_grid` rows so toggling visibility
        won't change overall row heights.
        """
        try:
            top = self.winfo_toplevel()
            top.update_idletasks()
            mg = getattr(self, "main_grid", None)
            if mg is None:
                return

            total = mg.winfo_height() or top.winfo_height() or 0
            for r in (0, 1, 2):
                max_h = 0
                for child in mg.grid_slaves(row=r):
                    try:
                        h = child.winfo_reqheight() or child.winfo_height() or 0
                    except Exception:
                        h = 0
                    if h > max_h:
                        max_h = h
                if max_h <= 0:
                    max_h = max(120, (total // 3) if total else 160)
                try:
                    mg.rowconfigure(r, minsize=max_h + 6)
                except Exception:
                    pass
            try:
                mg.grid_propagate(False)
            except Exception:
                pass
            top.update_idletasks()
        except Exception:
            pass

    def _ctx_new(self):
        name = askstring("Nowy kontekst", "Podaj nazwę szkoły/placówki (np. „SP Górzno”):")
        if not name:
            return
        name = name.strip()
        if not name:
            return
        if name in self.cfg["contexts"]:
            messagebox.showerror("Kontekst", "Taki kontekst już istnieje.")
            return
        self.cfg["contexts"][name] = _new_ctx_defaults()
        self.cfg["current_context"] = name
        save_cfg(self.cfg)
        self.ctx_name.set(name)
        self._reload_ctx_list()
        self._reload_from_context()

    def _ctx_manage(self):
        win = tk.Toplevel(self.winfo_toplevel())
        win.title("Zarządzanie kontekstem")
        pad = 10
        frm = ttk.Frame(win, padding=pad)
        frm.pack(fill="both", expand=True)
        ttk.Label(frm, text=f"Aktualny: {self.ctx_name.get()}", style="Flat.TLabel").grid(
            row=0, column=0, sticky="w", columnspan=2
        )
        ttk.Button(frm, text="Zmień nazwę…", command=lambda: self._ctx_rename(win), style="TButton").grid(
            row=1, column=0, sticky="w", pady=(8, 0)
        )
        ttk.Button(frm, text="Usuń…", command=lambda: self._ctx_delete(win), style="TButton").grid(
            row=1, column=1, sticky="w", pady=(8, 0)
        )

    def _ctx_rename(self, owner):
        old = self.ctx_name.get()
        new = askstring("Zmień nazwę", "Nowa nazwa kontekstu:", parent=owner)
        if not new:
            return
        new = new.strip()
        if not new:
            return
        try:
            rename_ctx(self.cfg, old, new)
        except Exception as e:
            messagebox.showerror("Kontekst", str(e))
            return
        self.ctx_name.set(new)
        self._reload_ctx_list()
        self._reload_from_context()

    def _ctx_delete(self, owner):
        name = self.ctx_name.get()
        if messagebox.askyesno(
            "Usuń kontekst",
            f"Czy na pewno usunąć „{name}”?\n(Ustawienia tej szkoły zostaną usunięte)",
        ):
            try:
                delete_ctx(self.cfg, name)
            except Exception as e:
                messagebox.showerror("Kontekst", str(e))
                return
            self.ctx_name.set(get_current_ctx_name(self.cfg))
            self._reload_ctx_list()
            self._reload_from_context()

    def _reload_ctx_list(self):
        self.cb_ctx.configure(values=list(self.cfg["contexts"].keys()))

    def _reload_from_context(self):
        switch_ctx(self.cfg, self.ctx_name.get())
        ctx = get_ctx(self.cfg)
        self.file_path.set(ctx.get("last_file", ""))
        self.max_points.set(str(ctx.get("max_points", 60)))
        self.open_after.set(bool(ctx.get("open_after", True)))
        self.output_dir.set(ctx.get("last_output_dir", ""))
        self.use_weighted.set(bool(ctx.get("use_weighted_mean", False)))
        self.round_before.set(bool(ctx.get("round_percent_before_grade", False)))
        label = ctx.get("scale_active", "Domyślna")
        self.active_scale_name.set(label)
        profiles = set(SCALE_PROFILES.keys())
        profiles |= set((ctx.get("custom_scales") or {}).keys())
        profiles = list(sorted(profiles))
        if label not in profiles:
            profiles.insert(0, label)
        self.cb_profile.configure(values=profiles)

    def _on_profile_selected(self, _e=None):
        name = self.active_scale_name.get()
        ctx = get_ctx(self.cfg).copy()
        base = (ctx.get("custom_scales") or {}).get(name) or SCALE_PROFILES.get(name)
        if base:
            ctx["active_scale_rows"] = [(int(a), int(b), str(lbl)) for a, b, lbl in base]
            ctx["scale_active"] = name
            set_ctx(self.cfg, get_current_ctx_name(self.cfg), ctx)
            messagebox.showinfo(
                "Skala",
                f"Ustawiono profil „{name}” dla kontekstu „{get_current_ctx_name(self.cfg)}”.",
            )

    # ----- Edycja wag arkuszy
    def edit_weights(self):
        in_path = self.file_path.get().strip()
        if not in_path or not Path(in_path).exists():
            messagebox.showwarning(APP_TITLE, "Najpierw wybierz plik (.xlsx/.xls/.ods/.csv), aby pobrać listę arkuszy.")
            return
        try:
            sheet_names = list(read_input_frames(in_path).keys())
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Nie mogę odczytać arkuszy:\n{e}")
            return

        ctx = get_ctx(self.cfg).copy()
        weights = dict(ctx.get("weights_by_sheet") or {})

        win = tk.Toplevel(self.winfo_toplevel())
        win.title(f"Wagi arkuszy – {self.ctx_name.get()}")
        win.resizable(False, False)
        pad = 10
        frm = ttk.Frame(win, padding=pad)
        frm.pack(fill="both", expand=True)
        ttk.Label(frm, text="Ustaw wagi dla arkuszy (domyślnie 1.0).", style="Hint.TLabel").grid(
            row=0, column=0, columnspan=2, sticky="w"
        )
        entries = {}
        r = 1
        for nm in sheet_names:
            ttk.Label(frm, text=nm, style="Flat.TLabel").grid(row=r, column=0, sticky="w", padx=(0, 10))
            e = ttk.Entry(frm, width=8)
            e.insert(0, str(weights.get(nm, 1.0)))
            e.grid(row=r, column=1, sticky="w")
            entries[nm] = e
            r += 1

        def save():
            try:
                new = {nm: float(e.get()) for nm, e in entries.items()}
            except Exception:
                messagebox.showerror("Wagi", "Wagi muszą być liczbami.")
                return
            ctx2 = get_ctx(self.cfg).copy()
            ctx2["weights_by_sheet"] = new
            set_ctx(self.cfg, get_current_ctx_name(self.cfg), ctx2)
            messagebox.showinfo("Wagi", "Zapisano wagi.")
            win.destroy()

        ttk.Button(
            frm,
            text="Domyślne (1.0)",
            command=lambda: [e.delete(0, "end") or e.insert(0, "1.0") for e in entries.values()],
            style="TButton",
        ).grid(row=r, column=0, sticky="w", pady=(8, 0))
        ttk.Button(frm, text="Zapisz", command=save, style="Accent.TButton").grid(
            row=r, column=1, sticky="e", pady=(8, 0)
        )

    def edit_weight_profiles(self):
        """Zarządzanie profilami wag dla arkuszy w bieżącym kontekście."""
        in_path = self.file_path.get().strip()
        if not in_path or not Path(in_path).exists():
            messagebox.showwarning(
                APP_TITLE,
                "Najpierw wybierz plik (.xlsx/.xls/.ods/.csv), aby pobrać listę arkuszy."
            )
            return
        try:
            sheet_names = list(read_input_frames(in_path).keys())
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Nie mogę odczytać arkuszy:\n{e}")
            return

        ctx = get_ctx(self.cfg).copy()
        profiles = dict(ctx.get("weight_profiles") or {})
        active_name = ctx.get("active_weight_profile", "") or ""

        win = tk.Toplevel(self.winfo_toplevel())
        win.title(f"Profile wag – {self.ctx_name.get()}")
        win.resizable(False, False)
        pad = 10

        container = ttk.Frame(win, padding=pad)
        container.pack(fill="both", expand=True)

        # lewy panel: lista profili
        left = ttk.Frame(container)
        left.grid(row=0, column=0, sticky="nsw", padx=(0, 10))
        ttk.Label(left, text="Profile wag:", style="Flat.TLabel").pack(anchor="w")

        listbox = tk.Listbox(left, height=8)
        listbox.pack(fill="y", expand=False, pady=(4, 0))
        sb = ttk.Scrollbar(left, orient="vertical", command=listbox.yview)
        listbox.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")

        # wypełnij listę nazw profili
        for name in sorted(profiles.keys(), key=str.lower):
            listbox.insert("end", name)
            if name == active_name:
                listbox.selection_set("end")

        # prawy panel: szczegóły profilu
        right = ttk.Frame(container)
        right.grid(row=0, column=1, sticky="nsew")
        container.columnconfigure(1, weight=1)

        ttk.Label(right, text="Nazwa profilu:", style="Flat.TLabel").grid(row=0, column=0, sticky="w")
        profile_name_var = tk.StringVar(value=active_name)
        name_entry = ttk.Entry(right, textvariable=profile_name_var, width=40, style="Flat.TEntry")
        name_entry.grid(row=0, column=1, sticky="we", padx=(8, 0))

        ttk.Label(
            right,
            text="Wagi dla arkuszy (domyślnie 1.0):",
            style="Hint.TLabel",
        ).grid(row=1, column=0, columnspan=2, sticky="w", pady=(6, 0))

        entries = {}
        r = 2
        for nm in sheet_names:
            ttk.Label(right, text=nm, style="Flat.TLabel").grid(row=r, column=0, sticky="w", padx=(0, 10))
            e = ttk.Entry(right, width=8, style="Flat.TEntry")
            e.insert(0, "1.0")
            e.grid(row=r, column=1, sticky="w")
            entries[nm] = e
            r += 1

        def load_profile(name: str):
            """Załaduj wagi z wybranego profilu do pól edycyjnych."""
            nonlocal profiles
            profile_name_var.set(name)
            data = profiles.get(name, {}) or {}
            for nm, entry in entries.items():
                entry.delete(0, "end")
                entry.insert(0, str(data.get(nm, 1.0)))

        def on_select(evt=None):
            sel = listbox.curselection()
            if not sel:
                return
            name = listbox.get(sel[0])
            load_profile(name)

        listbox.bind("<<ListboxSelect>>", on_select)

        # jeśli jest aktywny profil – załaduj jego dane
        if active_name and active_name in profiles:
            load_profile(active_name)

        btn_frame = ttk.Frame(right)
        btn_frame.grid(row=r, column=0, columnspan=2, sticky="e", pady=(10, 0))

        def new_profile():
            profile_name_var.set("")
            for e in entries.values():
                e.delete(0, "end")
                e.insert(0, "1.0")
            listbox.selection_clear(0, "end")

        def save_profile():
            nonlocal profiles
            name = profile_name_var.get().strip()
            if not name:
                messagebox.showerror("Profile wag", "Podaj nazwę profilu.")
                return
            try:
                weights = {nm: float(e.get()) for nm, e in entries.items()}
            except Exception:
                messagebox.showerror("Profile wag", "Wagi muszą być liczbami.")
                return
            ctx2 = get_ctx(self.cfg).copy()
            profiles = dict(ctx2.get("weight_profiles") or {})
            profiles[name] = weights
            ctx2["weight_profiles"] = profiles
            set_ctx(self.cfg, get_current_ctx_name(self.cfg), ctx2)

            # odśwież listę
            listbox.delete(0, "end")
            for n in sorted(profiles.keys(), key=str.lower):
                listbox.insert("end", n)
            # zaznacz zapisywany profil
            for idx in range(listbox.size()):
                if listbox.get(idx) == name:
                    listbox.selection_clear(0, "end")
                    listbox.selection_set(idx)
                    break
            messagebox.showinfo("Profile wag", "Zapisano profil wag.")

        def set_active_profile():
            sel_name = profile_name_var.get().strip()
            if not sel_name:
                messagebox.showwarning("Profile wag", "Najpierw podaj nazwę profilu lub wybierz z listy.")
                return
            ctx2 = get_ctx(self.cfg).copy()
            profiles_loc = dict(ctx2.get("weight_profiles") or {})
            if sel_name not in profiles_loc:
                # spróbuj najpierw zapisać aktualny stan jako profil
                try:
                    weights = {nm: float(e.get()) for nm, e in entries.items()}
                except Exception:
                    messagebox.showerror("Profile wag", "Wagi muszą być liczbami.")
                    return
                profiles_loc[sel_name] = weights
                ctx2["weight_profiles"] = profiles_loc
            weights_for_active = profiles_loc[sel_name]
            ctx2["active_weight_profile"] = sel_name
            # Ustaw także bieżące weights_by_sheet na podstawie profilu
            ctx2["weights_by_sheet"] = dict(weights_for_active)
            set_ctx(self.cfg, get_current_ctx_name(self.cfg), ctx2)
            messagebox.showinfo("Profile wag", f"Ustawiono aktywny profil wag: {sel_name}")

        def delete_profile():
            nonlocal profiles
            sel = listbox.curselection()
            if not sel:
                messagebox.showwarning("Profile wag", "Najpierw wybierz profil z listy.")
                return
            name = listbox.get(sel[0])
            if not messagebox.askyesno("Profile wag", f"Czy na pewno usunąć profil: {name}?"):
                return
            ctx2 = get_ctx(self.cfg).copy()
            profiles = dict(ctx2.get("weight_profiles") or {})
            if name in profiles:
                profiles.pop(name, None)
                ctx2["weight_profiles"] = profiles
                if ctx2.get("active_weight_profile") == name:
                    ctx2["active_weight_profile"] = ""
                set_ctx(self.cfg, get_current_ctx_name(self.cfg), ctx2)
            # odśwież listę
            listbox.delete(0, "end")
            for n in sorted(profiles.keys(), key=str.lower):
                listbox.insert("end", n)
            new_profile()

        ttk.Button(btn_frame, text="Nowy profil", command=new_profile, style="TButton").pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(btn_frame, text="Zapisz profil", command=save_profile, style="Accent.TButton").pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(btn_frame, text="Ustaw jako aktywny", command=set_active_profile, style="TButton").pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(btn_frame, text="Usuń profil", command=delete_profile, style="TButton").pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(btn_frame, text="Zamknij", command=win.destroy, style="TButton").pack(
            side="left", padx=(12, 0)
        )


    # ----- Ręczne wprowadzanie danych (imię + punkty)
    def open_manual_input(self):
        win = tk.Toplevel(self.winfo_toplevel())
        win.title("Ręczne wprowadzanie wyników")
        win.transient(self.winfo_toplevel())
        win.grab_set()
        pad = 10

        frm = ttk.Frame(win, padding=pad, style="Flat.TFrame")
        frm.pack(fill="both", expand=True)

        ttk.Label(
            frm,
            text=(
                "Wpisz lub wklej dane uczniów.\n"
                "Możesz skorzystać z rubryk poniżej (Imię i nazwisko + liczba punktów, przycisk „Dodaj do listy”),\n"
                "albo wkleić gotową listę w polu tekstowym – każda linia w formacie: Imię i nazwisko;punkty.\n"
                "Dopuszczalne są też separatory TAB lub ostatnia spacja."
            ),
            style="Flat.TLabel",
            justify="left",
        ).grid(row=0, column=0, columnspan=2, sticky="w")

        inputs = ttk.Frame(frm, style="Flat.TFrame")
        inputs.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(6, 0))

        ttk.Label(inputs, text="Imię i nazwisko:", style="Flat.TLabel").grid(row=0, column=0, sticky="w")
        entry_name = ttk.Entry(inputs, width=36, style="Flat.TEntry")
        entry_name.grid(row=0, column=1, sticky="w", padx=(4, 10))

        ttk.Label(inputs, text="Liczba punktów:", style="Flat.TLabel").grid(row=0, column=2, sticky="w")
        entry_pts = ttk.Entry(inputs, width=10, style="Flat.TEntry")
        entry_pts.grid(row=0, column=3, sticky="w", padx=(4, 0))

        text_frame = ttk.Frame(frm, style="Flat.TFrame")
        text_frame.grid(row=2, column=0, columnspan=2, sticky="nsew", pady=(6, 0))
        frm.rowconfigure(2, weight=1)
        frm.columnconfigure(0, weight=1)

        txt = tk.Text(text_frame, width=70, height=20, wrap="none")
        vsb = ttk.Scrollbar(text_frame, orient="vertical", command=txt.yview)
        txt.configure(yscrollcommand=vsb.set)
        txt.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        def add_line():
            name = entry_name.get().strip()
            pts_str = entry_pts.get().strip().replace(",", ".")
            if not name or not pts_str:
                messagebox.showwarning(APP_TITLE, "Uzupełnij zarówno imię i nazwisko, jak i liczbę punktów.")
                return
            try:
                float(pts_str)
            except Exception:
                messagebox.showerror(APP_TITLE, f"Liczba punktów („{pts_str}”) musi być liczbą.")
                return

            line = f"{name};{pts_str}"
            current = txt.get("1.0", "end").strip()
            if current:
                txt.insert("end", "\n" + line)
            else:
                txt.insert("end", line)

            entry_name.delete(0, "end")
            entry_pts.delete(0, "end")
            entry_name.focus_set()

        ttk.Button(inputs, text="Dodaj do listy", command=add_line, style="TButton").grid(
            row=0, column=4, sticky="w", padx=(10, 0)
        )
        entry_pts.bind("<Return>", lambda e: add_line())

        btns = ttk.Frame(frm, style="Flat.TFrame")
        btns.grid(row=3, column=0, columnspan=2, sticky="e", pady=(8, 0))

        def do_save():
            raw = txt.get("1.0", "end").strip()
            if not raw:
                messagebox.showwarning(
                    APP_TITLE,
                    "Wprowadź co najmniej jednego ucznia (rubryki lub pole tekstowe).",
                )
                return

            rows = []
            line_no = 0
            for line in raw.splitlines():
                line_no += 1
                line = line.strip()
                if not line:
                    continue

                name = None
                pts_str = None

                if ";" in line:
                    name, pts_str = line.rsplit(";", 1)
                elif "\t" in line:
                    name, pts_str = line.rsplit("\t", 1)
                else:
                    parts = line.rsplit(" ", 1)
                    if len(parts) != 2:
                        messagebox.showerror(
                            APP_TITLE,
                            f"Linia {line_no}: nie mogę odczytać punktów.\n"
                            "Użyj formatu: Imię i nazwisko;punkty",
                        )
                        return
                    name, pts_str = parts

                name = name.strip()
                pts_str = pts_str.strip().replace(",", ".")

                if not name:
                    messagebox.showerror(APP_TITLE, f"Linia {line_no}: brak imienia i nazwiska.")
                    return
                try:
                    pts = float(pts_str)
                except Exception:
                    messagebox.showerror(
                        APP_TITLE,
                        f"Linia {line_no}: punkty („{pts_str}”) nie są liczbą.",
                    )
                    return

                rows.append((name, pts))

            if not rows:
                messagebox.showwarning(APP_TITLE, "Brak poprawnych danych.")
                return

            try:
                max_points = float(self.max_points.get().strip().replace(",", "."))
                if max_points <= 0:
                    raise ValueError()
            except Exception:
                messagebox.showerror(APP_TITLE, "Podaj poprawną maksymalną liczbę punktów (np. 60).")
                return

            df = pd.DataFrame(rows, columns=["Nazwisko", "Ilość punktów"])
            scale_rows = active_scale_from_ctx(self.cfg)
            round_before = bool(get_ctx(self.cfg).get("round_percent_before_grade", False))

            default_name = "wyniki_reczne_przetworzone.xlsx"
            out_path = filedialog.asksaveasfilename(
                title="Zapisz wynik jako…",
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel XLSX", "*.xlsx")],
            )
            if not out_path:
                return

            self.status.set("Przetwarzanie danych z formularza…")
            self.progress["value"] = 0
            self.progress["maximum"] = 1

            try:
                df_out = sanitize_and_recompute(df, max_points, scale_rows, round_before)
                write_multi_with_formatting(
                    {"Wyniki": df_out},
                    out_path,
                    use_weighted=False,
                    weights_by_sheet={},
                    scale_rows=scale_rows,
                    round_before=round_before,
                    max_points=max_points,
                )

                meta = {
                    "source": "manual_input",
                    "max_points": max_points,
                    "round_before": round_before,
                    "out_path": out_path,
                    "scale_rows": scale_rows,
                    "class_name": self.class_name.get().strip(),
                    "subject": (self.subject_var.get().strip() or ""),
                    "school": (self.school_var.get().strip() or self.ctx_name.get().strip() or ""),
                }
                title = f"Ręczne wprowadzanie – {Path(out_path).stem}"
                try:
                    archive_path = save_result_to_archive(self.ctx_name.get(), title, df_out, meta)
                    self._last_archive_path = archive_path
                except Exception:
                    archive_path = None

            except PermissionError:
                messagebox.showerror(
                    APP_TITLE,
                    "Nie można zapisać pliku. Zamknij go w Excelu i spróbuj ponownie.",
                )
                self.status.set("Błąd zapisu – plik zajęty?")
                return
            except Exception as e:
                tb = traceback.format_exc()
                messagebox.showerror(
                    APP_TITLE,
                    f"Wystąpił błąd:\n{e}\n\nSzczegóły:\n{tb}",
                )
                self.status.set("Błąd.")
                return
            else:
                msg = f"Kontekst: {self.ctx_name.get()}\nGotowe!\nZapisano:\n{out_path}"
                messagebox.showinfo(APP_TITLE, msg)
                self.status.set("Zakończono pomyślnie.")
                if self.open_after.get():
                    try:
                        os.startfile(out_path)
                    except Exception:
                        pass

                if archive_path is not None:
                    open_archive_window(self.winfo_toplevel(), preselect=archive_path)

            finally:
                self.progress["value"] = 1

            win.destroy()

        ttk.Button(btns, text="Zapisz do Excela", command=do_save, style="Accent.TButton").pack(side="right")
        ttk.Button(btns, text="Anuluj", command=win.destroy, style="TButton").pack(side="right", padx=(8, 0))

    # ----- Archiwum – przyciski
    def open_archive(self):
        open_archive_window(self.winfo_toplevel())

    def show_last_result(self):
        if not self._last_archive_path:
            messagebox.showinfo(
                APP_TITLE,
                "Brak zapamiętanego wyniku.\nNajpierw przelicz sprawdzian lub wprowadź dane ręcznie.",
            )
            return
        path = Path(self._last_archive_path)
        if not path.exists():
            messagebox.showwarning(
                APP_TITLE,
                "Ostatni zapamiętany wynik nie istnieje już na dysku (plik archiwum został usunięty).",
            )
            self._last_archive_path = None
            return
        open_archive_window(self.winfo_toplevel(), preselect=path)

    # ----- Batch toggle
    def _toggle_batch(self, init=False):
        # Use explicit spacing (pad was local to __init__); values assume pad=10
        if init:
            # Initial setup: ensure hidden/shown state without changing user-visible widgets unnecessarily
            if self.batch_mode.get():
                # batch mode active at start -> show batch controls
                try:
                    self.lf_file._outer.grid_remove()
                    self.lf_files._outer.grid(row=1, column=1, sticky="nsew", padx=(5, 0), pady=(3, 3))
                except Exception:
                    pass
                try:
                    # show output folder under batch files (row=2, col=1)
                    self.lf_out._outer.grid(row=2, column=1, sticky="nsew", padx=(5, 0), pady=(3, 0))
                except Exception:
                    pass
                # remove any spacer placeholders
                try:
                    self._spacer_r1.grid_remove()
                except Exception:
                    pass
                try:
                    self._spacer_r2.grid_remove()
                except Exception:
                    pass
            else:
                # single mode at start -> ensure batch controls hidden
                try:
                    self.lf_files._outer.grid_remove()
                except Exception:
                    pass
                try:
                    self.lf_file._outer.grid(row=1, column=1, sticky="nsew", padx=(5, 0), pady=(3, 3))
                except Exception:
                    pass
                try:
                    self.lf_out._outer.grid_remove()
                except Exception:
                    pass
                # show spacer in row=2 to keep row heights stable
                try:
                    self._spacer_r2.grid(row=2, column=1, sticky="nsew", padx=(5, 0), pady=(3, 0))
                except Exception:
                    pass
            return

        # Normal toggle (user interaction)
        if self.batch_mode.get():
            # Switch to batch: hide single file, show batch files and output folder
            self.lf_file._outer.grid_remove()
            self.lf_files._outer.grid(row=1, column=1, sticky="nsew", padx=(5, 0), pady=(3, 3))
            try:
                self.lf_out._outer.grid(row=2, column=1, sticky="nsew", padx=(5, 0), pady=(3, 0))
            except Exception:
                pass
            # remove placeholders so real widgets determine heights
            try:
                self._spacer_r1.grid_remove()
            except Exception:
                pass
            try:
                self._spacer_r2.grid_remove()
            except Exception:
                pass
            if not init:
                self.status.set("Tryb wsadowy: wybierz pliki i folder wyjściowy.")
        else:
            # Switch to single: hide batch files and output folder, show single file selector
            self.lf_files._outer.grid_remove()
            self.lf_file._outer.grid(row=1, column=1, sticky="nsew", padx=(5, 0), pady=(3, 3))
            try:
                self.lf_out._outer.grid_remove()
            except Exception:
                pass
            # ensure spacer occupies row=2 so heights don't shift
            try:
                self._spacer_r2.grid(row=2, column=1, sticky="nsew", padx=(5, 0), pady=(3, 0))
            except Exception:
                pass
            if not init:
                self.status.set("Tryb pojedynczy: wybierz plik.")

    # ----- DnD
    def _on_drop_single(self, event):
        raw = event.data.strip()
        if raw.startswith("{") and raw.endswith("}"):
            raw = raw[1:-1]
        path = raw.split("} {")[0]
        if path.lower().endswith((".xlsx", ".xls", ".ods", ".csv")):
            self.file_path.set(path)
            ctx = get_ctx(self.cfg).copy()
            ctx["last_file"] = path
            set_ctx(self.cfg, get_current_ctx_name(self.cfg), ctx)
        else:
            messagebox.showwarning(APP_TITLE, "Upuść plik (.xlsx/.xls/.ods/.csv).")

    def _on_drop_multi(self, event):
        data = event.data.strip()
        files = []
        if data.startswith("{") and data.endswith("}"):
            for p in data.split("} {"):
                p = p.strip("{}")
                if p.lower().endswith((".xlsx", ".xls", ".ods", ".csv")):
                    files.append(p)
        else:
            if data.lower().endswith((".xlsx", ".xls", ".ods", ".csv")):
                files.append(data)
        if files:
            self.batch_files = files
            self.lbl_files.configure(text=f"Wybrano: {len(files)} plików")
        else:
            messagebox.showwarning(APP_TITLE, "Upuść pliki (.xlsx/.xls/.ods/.csv).")

    # ----- pickers
    def pick_file(self):
        cfg = self.cfg if isinstance(self.cfg, dict) else load_cfg()
        initial_dir = ""
        try:
            if cfg.get("remember_last_dir", True):
                initial_dir = cfg.get("last_input_dir", "")
        except Exception:
            initial_dir = ""

        dialog_kwargs = {
            "title": "Wybierz plik (.xlsx/.xls/.ods/.csv)",
            "filetypes": [
                ("Arkusze Excel/Calc/CSV", "*.xlsx;*.xls;*.ods;*.csv"),
                ("Excel XLSX", "*.xlsx"),
                ("Excel XLS", "*.xls"),
                ("LibreOffice ODS", "*.ods"),
                ("CSV", "*.csv"),
                ("Wszystkie pliki", "*.*"),
            ],
        }
        if initial_dir:
            dialog_kwargs["initialdir"] = initial_dir

        p = filedialog.askopenfilename(**dialog_kwargs)
        if p:
            self.file_path.set(p)
            ctx = get_ctx(self.cfg).copy()
            ctx["last_file"] = p
            set_ctx(self.cfg, get_current_ctx_name(self.cfg), ctx)
            try:
                if cfg.get("remember_last_dir", True):
                    cfg["last_input_dir"] = str(Path(p).parent)
                    self.cfg = cfg
                    save_cfg(cfg)
            except Exception:
                pass

    def pick_files(self):
        cfg = self.cfg if isinstance(self.cfg, dict) else load_cfg()
        initial_dir = ""
        try:
            if cfg.get("remember_last_dir", True):
                initial_dir = cfg.get("last_input_dir", "")
        except Exception:
            initial_dir = ""

        dialog_kwargs = {
            "title": "Wybierz pliki (.xlsx/.xls/.ods/.csv) – wiele",
            "filetypes": [
                ("Arkusze Excel/Calc/CSV", "*.xlsx;*.xls;*.ods;*.csv"),
                ("Excel XLSX", "*.xlsx"),
                ("Excel XLS", "*.xls"),
                ("LibreOffice ODS", "*.ods"),
                ("CSV", "*.csv"),
            ],
        }
        if initial_dir:
            dialog_kwargs["initialdir"] = initial_dir

        files = filedialog.askopenfilenames(**dialog_kwargs)
        if files:
            self.batch_files = list(files)
            self.lbl_files.configure(text=f"Wybrano: {len(files)} plików")
            try:
                if cfg.get("remember_last_dir", True):
                    cfg["last_input_dir"] = str(Path(self.batch_files[0]).parent)
                    self.cfg = cfg
                    save_cfg(cfg)
            except Exception:
                pass

    def pick_output_dir(self):
        d = filedialog.askdirectory(title="Wybierz folder wyjściowy")
        if d:
            self.output_dir.set(d)
            ctx = get_ctx(self.cfg).copy()
            ctx["last_output_dir"] = d
            set_ctx(self.cfg, get_current_ctx_name(self.cfg), ctx)

    def open_settings_window(self):
        """
        Okno ustawień programu:
        - domyślna szkoła / przedmiot / klasa,
        - pamiętanie ostatniego folderu,
        - wybór motywu interfejsu (jasny / ciemny).
        """
        cfg = _ensure_cfg_structure(self.cfg if isinstance(self.cfg, dict) else load_cfg())
        self.cfg = cfg

        default_school = cfg.get("default_school", "")
        default_subject = cfg.get("default_subject", "")
        default_class = cfg.get("default_class", "")
        remember_last_dir = cfg.get("remember_last_dir", True)
        current_theme = cfg.get("ui_theme", "light")
        if current_theme not in ("light", "dark"):
            current_theme = "light"

        win = tk.Toplevel(self.master)
        win.title("Ustawienia programu")
        win.transient(self.master)
        win.grab_set()
        win.minsize(460, 320)

        frm = ttk.Frame(win, padding=10, style="Flat.TFrame")
        frm.pack(fill="both", expand=True)

        # Zmienne powiązane z polami
        school_var = tk.StringVar(value=default_school)
        subject_var = tk.StringVar(value=default_subject)
        class_var = tk.StringVar(value=default_class)
        remember_var = tk.BooleanVar(value=remember_last_dir)
        theme_var = tk.StringVar(value=current_theme)

        row = 0
        ttk.Label(frm, text="Domyślna szkoła:", style="Flat.TLabel").grid(row=row, column=0, sticky="w")
        ttk.Entry(frm, textvariable=school_var, width=40, style="Flat.TEntry").grid(
            row=row, column=1, sticky="we", padx=(8, 0)
        )
        row += 1

        ttk.Label(frm, text="Domyślny przedmiot:", style="Flat.TLabel").grid(
            row=row, column=0, sticky="w", pady=(6, 0)
        )
        ttk.Entry(frm, textvariable=subject_var, width=40, style="Flat.TEntry").grid(
            row=row, column=1, sticky="we", padx=(8, 0), pady=(6, 0)
        )
        row += 1

        ttk.Label(frm, text="Domyślna klasa / grupa:", style="Flat.TLabel").grid(
            row=row, column=0, sticky="w", pady=(6, 0)
        )
        ttk.Entry(frm, textvariable=class_var, width=40, style="Flat.TEntry").grid(
            row=row, column=1, sticky="we", padx=(8, 0), pady=(6, 0)
        )
        row += 1

        ttk.Checkbutton(
            frm,
            text="Pamiętaj ostatni użyty folder wyjściowy",
            variable=remember_var,
            style="Flat.TCheckbutton",
        ).grid(row=row, column=0, columnspan=2, sticky="w", pady=(8, 0))
        row += 1

        # Wybór motywu interfejsu
        lf_theme = ttk.LabelFrame(frm, text="Motyw interfejsu", padding=(8, 6), style="Flat.TLabelframe")
        lf_theme.grid(row=row, column=0, columnspan=2, sticky="we", pady=(10, 0))
        lf_theme.columnconfigure(0, weight=1)

        ttk.Radiobutton(
            lf_theme,
            text="Jasny (nowoczesny)",
            value="light",
            variable=theme_var,
            style="Flat.TRadiobutton",
        ).grid(row=0, column=0, sticky="w")
        ttk.Radiobutton(
            lf_theme,
            text="Ciemny (nowoczesny)",
            value="dark",
            variable=theme_var,
            style="Flat.TRadiobutton",
        ).grid(row=1, column=0, sticky="w")
        row += 1

        # Przyciski
        btn_frame = ttk.Frame(frm, style="Flat.TFrame")
        btn_frame.grid(row=row, column=0, columnspan=2, sticky="e", pady=(12, 0))

        def on_cancel():
            win.destroy()

        def on_save():
            cfg["default_school"] = school_var.get().strip()
            cfg["default_subject"] = subject_var.get().strip()
            cfg["default_class"] = class_var.get().strip()
            cfg["remember_last_dir"] = bool(remember_var.get())
            cfg["ui_theme"] = theme_var.get()

            save_cfg(cfg)
            self.cfg = cfg

            # zaktualizuj bieżące pola w GUI
            try:
                self.school_var.set(cfg.get("default_school", ""))
            except Exception:
                pass
            try:
                self.subject_var.set(cfg.get("default_subject", ""))
            except Exception:
                pass
            try:
                self.class_name.set(cfg.get("default_class", ""))
            except Exception:
                pass

            # zmiana motywu na żywo
            self.ui_theme = cfg.get("ui_theme", "light")
            self._apply_theme(self.winfo_toplevel())

            win.destroy()

        ttk.Button(btn_frame, text="Anuluj", command=on_cancel, style="TButton").pack(side="right", padx=(8, 0))
        ttk.Button(btn_frame, text="Zapisz", command=on_save, style="Accent.TButton").pack(side="right")

    def open_manual(self):
        if Path(PATH_PDF).exists():
            try:
                os.startfile(str(PATH_PDF))
            except Exception as e:
                messagebox.showerror(APP_TITLE, f"Nie udało się otworzyć instrukcji.\n{e}")
        else:
            messagebox.showwarning(APP_TITLE, "Nie znaleziono Instrukcja_Wyniki5.pdf obok programu.")

    # ----- edytor skali
    def open_scale_editor(self):
        win = tk.Toplevel(self.winfo_toplevel())
        win.title(f"Skala ocen – {self.ctx_name.get()}")
        win.resizable(False, False)
        pad = 10
        frame = ttk.Frame(win, padding=pad)
        frame.pack(fill="both", expand=True)

        ctx = get_ctx(self.cfg)
        ttk.Label(frame, text="Profil (szablon do edycji):", style="Flat.TLabel").grid(row=0, column=0, sticky="w")
        profiles = set(SCALE_PROFILES.keys()) | set((ctx.get("custom_scales") or {}).keys())
        profiles = list(sorted(profiles))
        current_label = self.active_scale_name.get() if self.active_scale_name.get() in profiles else "Domyślna"
        prof_var = tk.StringVar(value=current_label)
        cb = ttk.Combobox(frame, state="readonly", values=profiles, textvariable=prof_var, width=30)
        cb.grid(row=0, column=1, sticky="w", padx=(6, 0))

        ttk.Label(frame, text="Od (%)", style="Flat.TLabel").grid(row=1, column=0, sticky="w")
        ttk.Label(frame, text="Do (%)", style="Flat.TLabel").grid(row=1, column=1, sticky="w")
        ttk.Label(frame, text="Etykieta", style="Flat.TLabel").grid(row=1, column=2, sticky="w", padx=(6, 0))
        rows = []

        def load_into_entries(profile_name: str):
            base = (ctx.get("custom_scales") or {}).get(profile_name) or SCALE_PROFILES.get(
                profile_name, DEFAULT_SCALE
            )
            for e in rows:
                for w in e:
                    w.destroy()
            rows.clear()
            irow = 2
            for lo, hi, lbl in base:
                e1 = ttk.Entry(frame, width=6)
                e1.insert(0, str(int(lo)))
                e1.grid(row=irow, column=0, padx=(0, 6), pady=2, sticky="w")
                e2 = ttk.Entry(frame, width=6)
                e2.insert(0, str(int(hi)))
                e2.grid(row=irow, column=1, padx=(0, 6), pady=2, sticky="w")
                e3 = ttk.Entry(frame, width=28)
                e3.insert(0, lbl)
                e3.grid(row=irow, column=2, padx=(0, 0), pady=2, sticky="w")
                rows.append((e1, e2, e3))
                irow += 1

        load_into_entries(prof_var.get())
        cb.bind("<<ComboboxSelected>>", lambda _e=None: load_into_entries(prof_var.get()))

        btns = ttk.Frame(win, padding=(pad, pad, pad, pad))
        btns.pack(fill="x")

        def get_rows_from_ui():
            new = []
            for e1, e2, e3 in rows:
                try:
                    lo = int(e1.get())
                    hi = int(e2.get())
                    lbl = e3.get().strip()
                    if not (0 <= lo <= 100 and 0 <= hi <= 100 and lo <= hi) or not lbl:
                        raise ValueError
                    new.append((lo, hi, lbl))
                except Exception:
                    messagebox.showerror("Skala", "Sprawdź progi (0–100, lo<=hi) i etykiety.")
                    return None
            new.sort(key=lambda x: -x[0])
            return new

        def save_and_use():
            new_scale = get_rows_from_ui()
            if new_scale is None:
                return
            set_active_scale_rows(self.cfg, new_scale, label="Moja skala (auto)")
            self.active_scale_name.set("Moja skala (auto)")
            vals = list(self.cb_profile["values"])
            if "Moja skala (auto)" not in vals:
                vals.insert(0, "Moja skala (auto)")
                self.cb_profile.configure(values=vals)
            messagebox.showinfo(
                "Skala",
                "Zapisano. Ta skala będzie używana automatycznie (dla tego kontekstu).",
            )
            win.destroy()

        def save_as_new_profile():
            new_scale = get_rows_from_ui()
            if new_scale is None:
                return
            name = askstring(
                "Nowy profil",
                "Podaj nazwę profilu (np. „SP A – skala 60%”):",
                parent=win,
            )
            if not name:
                return
            name = name.strip()
            if not name:
                return
            ctx2 = get_ctx(self.cfg).copy()
            cs = ctx2.get("custom_scales") or {}
            cs[name] = [(int(a), int(b), str(lbl)) for a, b, lbl in new_scale]
            ctx2["custom_scales"] = cs
            ctx2["active_scale_rows"] = cs[name]
            ctx2["scale_active"] = name
            set_ctx(self.cfg, get_current_ctx_name(self.cfg), ctx2)
            self.active_scale_name.set(name)
            vals = list(self.cb_profile["values"])
            if name not in vals:
                vals.append(name)
            self.cb_profile.configure(values=vals)
            messagebox.showinfo(
                "Skala",
                f"Zapisano nowy profil „{name}” i ustawiono jako aktywny (dla tego kontekstu).",
            )
            win.destroy()

        ttk.Button(btns, text="Zapisz i użyj tej skali", command=save_and_use, style="Accent.TButton").pack(
            side="right"
        )
        ttk.Button(
            btns,
            text="Zapisz jako nowy profil…",
            command=save_as_new_profile,
            style="TButton",
        ).pack(side="right", padx=(8, 0))
        ttk.Button(btns, text="Anuluj", command=win.destroy, style="TButton").pack(
            side="right", padx=(8, 0)
        )

    # ----- RUN
    def run(self):
        ctx = get_ctx(self.cfg).copy()
        try:
            max_points = float(self.max_points.get().strip().replace(",", "."))
            if max_points <= 0:
                raise ValueError()
        except Exception:
            messagebox.showerror(APP_TITLE, "Podaj poprawną maksymalną liczbę punktów (np. 60).")
            return
        ctx["max_points"] = max_points
        ctx["open_after"] = bool(self.open_after.get())
        ctx["last_output_dir"] = self.output_dir.get().strip()
        ctx["use_weighted_mean"] = bool(self.use_weighted.get())
        ctx["round_percent_before_grade"] = bool(self.round_before.get())
        set_ctx(self.cfg, get_current_ctx_name(self.cfg), ctx)

        scale_rows = active_scale_from_ctx(self.cfg)
        use_weighted = bool(get_ctx(self.cfg).get("use_weighted_mean", False))
        weights_map = dict(get_ctx(self.cfg).get("weights_by_sheet") or {})
        round_before = bool(get_ctx(self.cfg).get("round_percent_before_grade", False))

        self.btn_run.state(["disabled"])
        self.status.set("Przygotowanie do przetwarzania…")
        self.progress["value"] = 0
        self.progress["maximum"] = 1

        if self.batch_mode.get():
            if not self.batch_files:
                messagebox.showwarning(APP_TITLE, "Wybierz pliki .xlsx/.xls/.ods/.csv do przetworzenia.")
                self.btn_run.state(["!disabled"])
                self.status.set("Gotowy.")
                return
            out_dir = self.output_dir.get().strip()
            if not out_dir:
                messagebox.showwarning(APP_TITLE, "Wybierz folder wyjściowy (np. …\\wyniki).")
                self.btn_run.state(["!disabled"])
                self.status.set("Gotowy.")
                return
            Path(out_dir).mkdir(parents=True, exist_ok=True)
            thread = threading.Thread(
                target=self._run_batch_threaded,
                args=(max_points, scale_rows, use_weighted, weights_map, round_before, out_dir),
            )
            thread.start()
        else:
            in_path = self.file_path.get().strip()
            if not in_path:
                messagebox.showwarning(APP_TITLE, "Wybierz plik .xlsx/.xls/.ods/.csv.")
                self.btn_run.state(["!disabled"])
                self.status.set("Gotowy.")
                return
            ctx["last_file"] = in_path
            set_ctx(self.cfg, get_current_ctx_name(self.cfg), ctx)

            default_name = Path(in_path).stem + "_przetworzone.xlsx"
            out_path = filedialog.asksaveasfilename(
                title="Zapisz wynik jako…",
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[
                    ("Arkusze Excel/Calc/CSV", "*.xlsx;*.xls;*.ods;*.csv"),
                    ("Excel XLSX", "*.xlsx"),
                    ("Excel XLS", "*.xls"),
                    ("LibreOffice ODS", "*.ods"),
                    ("CSV", "*.csv"),
                ],
            )
            if not out_path:
                self.btn_run.state(["!disabled"])
                self.status.set("Gotowy.")
                return
            thread = threading.Thread(
                target=self._run_single_threaded,
                args=(in_path, max_points, out_path, scale_rows, use_weighted, weights_map, round_before),
            )
            thread.start()

    def _run_batch_threaded(self, max_points, scale_rows, use_weighted, weights_map, round_before, out_dir):
        total = len(self.batch_files)
        ok = 0
        fails = 0
        errors = []
        self.progress["maximum"] = total
        self.status.set("Przetwarzanie wsadowe…")

        last_archive_path: Path | None = None

        for i, f in enumerate(self.batch_files, start=1):
            try:
                base = Path(f).stem
                out_path = Path(out_dir) / (base + "_przetworzone.xlsx")
                result = process_file_all_sheets(
                    f, max_points, str(out_path), scale_rows, use_weighted, weights_map, round_before
                )
                ok += 1

                try:
                    if result:
                        first_name = next(iter(result.keys()))
                        df_for_archive = result[first_name]
                        meta = {
                            "source": "batch_file",
                            "input_path": f,
                            "output_path": str(out_path),
                            "sheet": first_name,
                            "max_points": max_points,
                            "round_before": round_before,
                            "use_weighted": use_weighted,
                            "scale_rows": scale_rows,
                            "sheet_weight": float(weights_map.get(first_name, 1.0)) if use_weighted else None,
                            "class_name": (self.class_name.get().strip() or first_name),
                        }
                        title = f"{Path(f).name} – {first_name}"
                        last_archive_path = save_result_to_archive(self.ctx_name.get(), title, df_for_archive, meta)
                except Exception:
                    pass

            except Exception as e:
                fails += 1
                errors.append(f"- {Path(f).name}: {e}")
            finally:
                self.progress["value"] = i
                self.progress.update_idletasks()
                self.status.set(f"Postęp: {i}/{total}")

        if last_archive_path is not None:
            self._last_archive_path = last_archive_path

        self.btn_run.state(["!disabled"])
        summary = (
            f"Kontekst: {self.ctx_name.get()}\nOK: {ok}\nBłędy: {fails}\nFolder wyjściowy:\n{out_dir}"
        )
        if errors:
            summary += "\n\nSzczegóły błędów:\n" + "\n".join(errors[:20])
            if len(errors) > 20:
                summary += f"\n…(+{len(errors) - 20} kolejnych)"
        messagebox.showinfo(APP_TITLE, summary)
        self.status.set("Zakończono wsadowo.")

    def _run_single_threaded(self, in_path, max_points, out_path, scale_rows, use_weighted, weights_map, round_before):
        self.status.set("Przetwarzanie…")
        self.progress["value"] = 0
        self.progress["maximum"] = 1

        archive_path: Path | None = None

        try:
            # Uzupełnij przedmiot / klasę / szkołę z arkusza META, jeśli nie podano ich w GUI
            try:
                subj_gui = (self.subject_var.get() or "").strip()
            except Exception:
                subj_gui = ""
            try:
                class_gui = (self.class_name.get() or "").strip()
            except Exception:
                class_gui = ""
            try:
                school_gui = (self.school_var.get() or "").strip()
            except Exception:
                school_gui = ""

            if (not subj_gui) or (not class_gui) or (not school_gui):
                try:
                    wb_meta = load_workbook(in_path, read_only=True, data_only=True)
                    meta_sheet_name = None
                    for cand in ("META", "_meta", "Meta", "meta"):
                        if cand in wb_meta.sheetnames:
                            meta_sheet_name = cand
                            break
                    if meta_sheet_name is not None:
                        ws_meta = wb_meta[meta_sheet_name]
                        found_subject = None
                        found_class = None
                        found_school = None
                        # szukamy w kilku pierwszych wierszach wpisów "Przedmiot", "Klasa / grupa" oraz "Szkoła"
                        for r in range(1, 11):
                            key = ws_meta.cell(row=r, column=1).value
                            if not isinstance(key, str):
                                continue
                            key_norm = key.strip().lower()
                            if key_norm == "przedmiot" and found_subject is None:
                                val = ws_meta.cell(row=r, column=2).value
                                if val is not None:
                                    found_subject = str(val).strip()
                            if key_norm in ("klasa", "klasa / grupa") and found_class is None:
                                val = ws_meta.cell(row=r, column=2).value
                                if val is not None:
                                    found_class = str(val).strip()
                            if key_norm in ("szkoła", "szkola", "szkoła / placówka", "szkola / placowka") and found_school is None:
                                val = ws_meta.cell(row=r, column=2).value
                                if val is not None:
                                    found_school = str(val).strip()
                        # Ustaw tylko te pola, które nie były ustawione w GUI
                        if found_subject and not subj_gui:
                            try:
                                self.subject_var.set(found_subject)
                            except Exception:
                                pass
                        if found_class and not class_gui:
                            try:
                                self.class_name.set(found_class)
                            except Exception:
                                pass
                        if found_school and not school_gui:
                            try:
                                self.school_var.set(found_school)
                            except Exception:
                                pass
                except Exception:
                    # brak pliku xlsx lub arkusza META – po prostu ignorujemy
                    pass

            result = process_file_all_sheets(
                in_path, max_points, out_path, scale_rows, use_weighted, weights_map, round_before
            )
        except PermissionError:
            messagebox.showerror(
                APP_TITLE,
                "Nie można zapisać pliku. Zamknij go w Excelu i spróbuj ponownie.",
            )
            self.status.set("Błąd zapisu – plik zajęty?")
        except Exception as e:
            tb = traceback.format_exc()
            messagebox.showerror(APP_TITLE, f"Wystąpił błąd:\n{e}\n\nSzczegóły:\n{tb}")
            self.status.set("Błąd.")
        else:
            msg = f"Kontekst: {self.ctx_name.get()}\nGotowe!\nZapisano:\n{out_path}"
            if use_weighted:
                msg += "\n(Uwzględniono średnią ważoną w zbiorczym podsumowaniu.)"
            messagebox.showinfo(APP_TITLE, msg)
            # krótkie podsumowanie wyników (pierwszy arkusz)
            summary_text = "Zakończono pomyślnie."
            try:
                if isinstance(result, dict) and result:
                    first_name_tmp = next(iter(result.keys()))
                    df_first = result.get(first_name_tmp)
                    if isinstance(df_first, pd.DataFrame) and not df_first.empty:
                        uczniowie = len(df_first)
                        avg_points = float(df_first["Ilość punktów"].mean())
                        avg_percent = float(df_first["Procent"].mean()) * 100.0
                        grade_counts = df_first["Ocena"].astype(str).str[0].value_counts()
                        dominant_grade = grade_counts.idxmax() if not grade_counts.empty else ""
                        avg_points_str = f"{avg_points:.1f}".replace(".", ",")
                        avg_percent_str = f"{avg_percent:.1f}".replace(".", ",")
                        summary_text = (
                            f"Uczniów: {uczniowie} | "
                            f"Średnia: {avg_points_str} pkt ({avg_percent_str}%) | "
                            f"Dominująca ocena: {dominant_grade}"
                        )
            except Exception:
                summary_text = "Zakończono pomyślnie."
            self.status.set(summary_text)
            if self.open_after.get():
                try:
                    os.startfile(out_path)
                except Exception:
                    pass

            try:
                if result:
                    first_name = next(iter(result.keys()))
                    df_for_archive = result[first_name]

                    # krótkie podsumowanie, które zapiszemy także w archiwum (pierwszy arkusz)
                    short_summary = None
                    try:
                        if isinstance(df_for_archive, pd.DataFrame) and not df_for_archive.empty:
                            uczniowie = len(df_for_archive)
                            avg_points = float(df_for_archive["Ilość punktów"].mean())
                            avg_percent = float(df_for_archive["Procent"].mean()) * 100.0
                            grade_counts = df_for_archive["Ocena"].astype(str).str[0].value_counts()
                            dominant_grade = grade_counts.idxmax() if not grade_counts.empty else ""
                            avg_points_str = f"{avg_points:.1f}".replace(".", ",")
                            avg_percent_str = f"{avg_percent:.1f}".replace(".", ",")
                            short_summary = (
                                f"Uczniów: {uczniowie} | "
                                f"Średnia: {avg_points_str} pkt ({avg_percent_str}%) | "
                                f"Dominująca ocena: {dominant_grade}"
                            )
                    except Exception:
                        short_summary = None

                    meta = {
                        "source": "single_file",
                        "input_path": in_path,
                        "output_path": out_path,
                        "sheet": first_name,
                        "max_points": max_points,
                        "round_before": round_before,
                        "use_weighted": use_weighted,
                        "scale_rows": scale_rows,
                        "sheet_weight": float(weights_map.get(first_name, 1.0)) if use_weighted else None,
                        "class_name": (self.class_name.get().strip() or first_name),
                        "subject": (self.subject_var.get().strip() or ""),
                        "school": (self.school_var.get().strip() or self.ctx_name.get().strip() or ""),
                        "short_summary": short_summary,
                    }
                    title = f"{Path(in_path).name} – {first_name}"
                    archive_path = save_result_to_archive(self.ctx_name.get(), title, df_for_archive, meta)
            except Exception:
                archive_path = None

            if archive_path is not None:
                self._last_archive_path = archive_path

        finally:
            self.progress["value"] = 1
            self.btn_run.state(["!disabled"])

    # ---------- konteksty – rename/delete ----------
    class ModernApp:
        """
        ULTRANOWOCZESNY interfejs oparty na customtkinter z zaawansowanym designem.
        Włącza gradient-like tła, nowoczesne kolory, animacje i responsywny layout.
        """

        def __init__(self):
            if not USE_CTK or ctk is None:
                raise RuntimeError("customtkinter nie jest zainstalowany")
            self.cfg = _ensure_cfg_structure(load_cfg())
            theme = self.cfg.get("ui_theme", "light")
            self.dark_mode = theme == "dark"
            ctk.set_appearance_mode("dark" if self.dark_mode else "light")
            try:
                ctk.set_default_color_theme("blue")
            except Exception:
                pass

            self.root = ctk.CTk()
            self.root.title(APP_TITLE + " — ULTRA")
            self.root.geometry("1300x800")
            self.root.minsize(1000, 600)

            # Kolory dla nowoczesnego designu
            self.primary_color = "#0066FF" if not self.dark_mode else "#0088FF"
            self.secondary_color = "#FF6B6B" if not self.dark_mode else "#FF8888"
            self.accent_color = "#FFD93D"
            self.bg_color = "#FFFFFF" if not self.dark_mode else "#1a1a1a"
            self.card_color = "#F8F9FF" if not self.dark_mode else "#2d2d2d"

            ctx = get_ctx(self.cfg)
            self.ctx_name = ctk.StringVar(value=get_current_ctx_name(self.cfg))
            self.school_var = ctk.StringVar(value=self.cfg.get("default_school", ""))
            self.subject_var = ctk.StringVar(value=self.cfg.get("default_subject", ""))
            self.class_name = ctk.StringVar(value=self.cfg.get("default_class", ""))
            self.file_path = ctk.StringVar(value=ctx.get("last_file", ""))
            self.max_points = ctk.StringVar(value=str(ctx.get("max_points", 60)))
            self.status = ctk.StringVar(value="Gotowy do przetwarzania")
            self.progress = None
            self.btn_run = None

            self._build_ui_ultra_modern()

        def _build_ui_ultra_modern(self):
            """Buduje ultra-nowoczesny interfejs z zaawansowanym designem."""
            # ===== TOP BAR (gradient-like) =====
            top_bar = ctk.CTkFrame(self.root, fg_color=self.primary_color, corner_radius=0, height=120)
            top_bar.pack(side="top", fill="x")
            top_bar.pack_propagate(False)

            # Logo and title
            logo_container = ctk.CTkFrame(top_bar, fg_color="transparent")
            logo_container.pack(side="left", padx=25, pady=20, anchor="w")

            ctk.CTkLabel(
                logo_container,
                text="⚡",
                font=("Segoe UI", 40, "bold"),
                text_color="white",
            ).pack(side="left", padx=(0, 12))

            title_container = ctk.CTkFrame(logo_container, fg_color="transparent")
            title_container.pack(side="left")
            ctk.CTkLabel(
                title_container,
                text="WYNIKI 5",
                font=("Segoe UI", 28, "bold"),
                text_color="white",
            ).pack(anchor="w")
            ctk.CTkLabel(
                title_container,
                text="System oceniania uczniów — Wersja ULTRA",
                font=("Segoe UI", 11),
                text_color="#B0D4FF",
            ).pack(anchor="w")

            # Top right buttons
            top_right = ctk.CTkFrame(top_bar, fg_color="transparent")
            top_right.pack(side="right", padx=25, pady=20, anchor="e")

            ctk.CTkButton(
                top_right,
                text="📋 Archiwum",
                command=self.open_archive,
                fg_color="#004CCC",
                hover_color="#0033AA",
                width=130,
                height=45,
                font=("Segoe UI", 11, "bold"),
            ).pack(side="left", padx=8)

            ctk.CTkButton(
                top_right,
                text="⚙️ Ustawienia",
                command=self._open_classic_settings,
                fg_color="#004CCC",
                hover_color="#0033AA",
                width=130,
                height=45,
                font=("Segoe UI", 11, "bold"),
            ).pack(side="left", padx=8)

            ctk.CTkButton(
                top_right,
                text="📖 Pomoc",
                command=self.open_manual,
                fg_color="#004CCC",
                hover_color="#0033AA",
                width=100,
                height=45,
                font=("Segoe UI", 11, "bold"),
            ).pack(side="left", padx=8)

            # ===== MAIN CONTENT =====
            main_container = ctk.CTkFrame(self.root, fg_color=self.bg_color)
            main_container.pack(fill="both", expand=True, padx=20, pady=20)
            main_container.columnconfigure(0, weight=1)
            main_container.columnconfigure(1, weight=2)

            # ===== LEFT PANEL - QUICK SETTINGS =====
            left_panel = ctk.CTkFrame(main_container, fg_color=self.card_color, corner_radius=15)
            left_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 15))

            # Left panel header
            left_header = ctk.CTkFrame(left_panel, fg_color="transparent")
            left_header.pack(fill="x", padx=20, pady=(20, 15))
            ctk.CTkLabel(
                left_header,
                text="⚙️ USTAWIENIA",
                font=("Segoe UI", 14, "bold"),
                text_color=self.primary_color,
            ).pack(anchor="w")

            # Settings fields
            settings_frame = ctk.CTkFrame(left_panel, fg_color="transparent")
            settings_frame.pack(fill="x", padx=20, pady=(0, 20))

            ctk.CTkLabel(settings_frame, text="Szkoła:", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(0, 4))
            ctk.CTkEntry(
                settings_frame,
                textvariable=self.school_var,
                placeholder_text="Nazwa szkoły/placówki",
                height=40,
            ).pack(fill="x", pady=(0, 10))

            ctk.CTkLabel(settings_frame, text="Przedmiot:", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(0, 4))
            ctk.CTkEntry(
                settings_frame,
                textvariable=self.subject_var,
                placeholder_text="np. Historia, Matematyka",
                height=40,
            ).pack(fill="x", pady=(0, 10))

            ctk.CTkLabel(settings_frame, text="Klasa/Grupa:", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(0, 4))
            ctk.CTkEntry(
                settings_frame,
                textvariable=self.class_name,
                placeholder_text="np. 6A, Grupa 1",
                height=40,
            ).pack(fill="x", pady=(0, 10))

            ctk.CTkLabel(settings_frame, text="Max. punktów:", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(0, 4))
            ctk.CTkEntry(
                settings_frame,
                textvariable=self.max_points,
                placeholder_text="60",
                height=40,
            ).pack(fill="x", pady=(0, 15))

            # Scale button
            ctk.CTkButton(
                settings_frame,
                text="🎚️ Edytor skali ocen",
                command=self._open_scale_editor,
                fg_color="#6366F1",
                hover_color="#4F46E5",
                height=35,
                font=("Segoe UI", 10, "bold"),
            ).pack(fill="x")

            # ===== RIGHT PANEL - FILE SELECTION & STATUS =====
            right_panel = ctk.CTkFrame(main_container, fg_color=self.card_color, corner_radius=15)
            right_panel.grid(row=0, column=1, sticky="nsew")
            right_panel.rowconfigure(0, weight=0)
            right_panel.rowconfigure(1, weight=1)
            right_panel.rowconfigure(2, weight=0)
            right_panel.columnconfigure(0, weight=1)

            # File section
            file_header = ctk.CTkFrame(right_panel, fg_color="transparent")
            file_header.pack(fill="x", padx=20, pady=(20, 15))
            ctk.CTkLabel(
                file_header,
                text="📁 PLIK Z WYNIKAMI",
                font=("Segoe UI", 14, "bold"),
                text_color=self.primary_color,
            ).pack(anchor="w")

            file_frame = ctk.CTkFrame(right_panel, fg_color="transparent")
            file_frame.pack(fill="x", padx=20, pady=(0, 10))

            file_entry_frame = ctk.CTkFrame(file_frame, fg_color="transparent")
            file_entry_frame.pack(fill="x")

            entry_file = ctk.CTkEntry(
                file_entry_frame,
                textvariable=self.file_path,
                placeholder_text="Wybierz plik .xlsx, .xls, .ods lub .csv...",
                height=40,
            )
            entry_file.pack(side="left", fill="both", expand=True, padx=(0, 10))

            ctk.CTkButton(
                file_entry_frame,
                text="📂",
                command=self.pick_file,
                fg_color=self.primary_color,
                hover_color="#004CCC",
                width=45,
                height=40,
                font=("Segoe UI", 18),
            ).pack(side="right")

            # Status section
            status_container = ctk.CTkFrame(right_panel, fg_color="transparent")
            status_container.pack(fill="both", expand=True, padx=20, pady=15)

            ctk.CTkLabel(
                status_container,
                text="📊 STATUS",
                font=("Segoe UI", 14, "bold"),
                text_color=self.primary_color,
            ).pack(anchor="w", pady=(0, 10))

            # Progress bar
            progress_label = ctk.CTkLabel(status_container, text="Postęp:", font=("Segoe UI", 10))
            progress_label.pack(anchor="w", pady=(0, 5))

            self.progress = ctk.CTkProgressBar(status_container, mode="determinate", height=8)
            self.progress.pack(fill="x", pady=(0, 12))
            self.progress.set(0)

            # Status text
            status_display = ctk.CTkLabel(
                status_container,
                textvariable=self.status,
                font=("Segoe UI", 11),
                justify="left",
                wraplength=400,
            )
            status_display.pack(fill="both", expand=True, padx=(0, 0))

            # Action buttons section
            action_container = ctk.CTkFrame(right_panel, fg_color="transparent")
            action_container.pack(fill="x", padx=20, pady=(15, 20))

            self.btn_run = ctk.CTkButton(
                action_container,
                text="▶️ PRZELICZ I ZAPISZ",
                command=self.run,
                fg_color="#00AA44",
                hover_color="#008833",
                height=55,
                font=("Segoe UI", 14, "bold"),
                text_color="white",
            )
            self.btn_run.pack(fill="x", pady=(0, 10))

            # Secondary button
            ctk.CTkButton(
                action_container,
                text="🔄 Tryb wsadowy",
                command=lambda: messagebox.showinfo(APP_TITLE, "Tryb wsadowy dostępny w interfejsie klasycznym."),
                fg_color="#FF6B6B",
                hover_color="#CC5555",
                height=35,
                font=("Segoe UI", 10, "bold"),
            ).pack(fill="x")

        def _open_classic_settings(self):
            # Użyj klasycznego okna ustawień z App (wymaga okna Tk). Otwieramy tymczasowe okno Tk.
            try:
                tmp = TkBase()
                tmp.withdraw()
                app = App(tmp)
                app.open_settings_window()
                tmp.destroy()
            except Exception:
                messagebox.showinfo(APP_TITLE, "Ustawienia dostępne w klasycznym interfejsie.")

        def pick_file(self):
            p = filedialog.askopenfilename(
                title="Wybierz plik (.xlsx/.xls/.ods/.csv)",
                filetypes=[("Arkusze Excel/CSV", "*.xlsx;*.xls;*.ods;*.csv"), ("Wszystkie pliki", "*.*")],
            )
            if p:
                self.file_path.set(p)
                ctx = get_ctx(self.cfg).copy()
                ctx["last_file"] = p
                set_ctx(self.cfg, get_current_ctx_name(self.cfg), ctx)

        def run(self):
            in_path = self.file_path.get().strip()
            if not in_path:
                messagebox.showwarning(APP_TITLE, "Wybierz plik .xlsx/.xls/.ods/.csv.")
                return
            try:
                max_points = float(self.max_points.get().strip().replace(",", "."))
                if max_points <= 0:
                    raise ValueError()
            except Exception:
                messagebox.showerror(APP_TITLE, "Podaj poprawną maksymalną liczbę punktów (np. 60).")
                return

            out_path = filedialog.asksaveasfilename(
                title="Zapisz wynik jako…",
                defaultextension=".xlsx",
                initialfile=Path(in_path).stem + "_przetworzone.xlsx",
                filetypes=[("Excel XLSX", "*.xlsx")],
            )
            if not out_path:
                return

            # uruchom w wątku
            thread = threading.Thread(target=self._run_thread, args=(in_path, max_points, out_path))
            thread.start()

        def _run_thread(self, in_path, max_points, out_path):
            self.status.set("Przetwarzanie…")
            self.progress.set(0)
            scale_rows = active_scale_from_ctx(self.cfg)
            use_weighted = bool(get_ctx(self.cfg).get("use_weighted_mean", False))
            weights_map = dict(get_ctx(self.cfg).get("weights_by_sheet") or {})
            round_before = bool(get_ctx(self.cfg).get("round_percent_before_grade", False))

            try:
                result = process_file_all_sheets(
                    in_path, max_points, out_path, scale_rows, use_weighted, weights_map, round_before
                )
            except Exception as e:
                tb = traceback.format_exc()
                messagebox.showerror(APP_TITLE, f"Błąd podczas przetwarzania:\n{e}\n\n{tb}")
                self.status.set("Błąd.")
                return

            # archiwizacja pierwszego arkusza
            try:
                if isinstance(result, dict) and result:
                    first_name = next(iter(result.keys()))
                    df_for_archive = result[first_name]
                    meta = {
                        "source": "modern_ui",
                        "input_path": in_path,
                        "output_path": out_path,
                        "sheet": first_name,
                        "max_points": max_points,
                    }
                    title = f"{Path(in_path).name} – {first_name}"
                    save_result_to_archive(get_current_ctx_name(self.cfg), title, df_for_archive, meta)
            except Exception:
                pass

            self.status.set(f"Zapisano: {out_path}")
            try:
                os.startfile(out_path)
            except Exception:
                pass

        def _open_scale_editor(self):
            """Otwiera edytor skali ocen."""
            try:
                tmp = TkBase()
                tmp.withdraw()
                app = App(tmp)
                app.open_scale_editor()
                tmp.destroy()
            except Exception as e:
                messagebox.showerror(APP_TITLE, f"Błąd otwarcia edytora skali:\n{e}")

        def open_archive(self):
            """Otwiera archiwum wyników."""
            try:
                ArchiveViewer(self.root)
            except Exception as e:
                messagebox.showerror(ARCHIVE_TITLE, f"Nie można otworzyć archiwum:\n{e}")

        def open_manual(self):
            """Otwiera instrukcję PDF."""
            if Path(PATH_PDF).exists():
                try:
                    os.startfile(str(PATH_PDF))
                except Exception:
                    messagebox.showinfo(APP_TITLE, "Instrukcja Wyniki5.pdf znajduje się w folderze programu.")
            else:
                messagebox.showwarning(APP_TITLE, "Nie znaleziono pliku Instrukcja_Wyniki5.pdf")

def rename_ctx(cfg, old, new):
    return _rename_ctx_impl(cfg, old, new)


def delete_ctx(cfg, name):
    return _delete_ctx_impl(cfg, name)


def _rename_ctx_impl(cfg, old, new):
    if old == new:
        return cfg
    if new in cfg["contexts"]:
        raise ValueError("Kontekst o takiej nazwie już istnieje.")
    cfg["contexts"][new] = cfg["contexts"].pop(old)
    if cfg["current_context"] == old:
        cfg["current_context"] = new
    save_cfg(cfg)
    return cfg


def _delete_ctx_impl(cfg, name):
    if name not in cfg["contexts"]:
        return cfg
    if len(cfg["contexts"]) == 1:
        raise ValueError("Musi pozostać co najmniej jeden kontekst.")
    del cfg["contexts"][name]
    if cfg["current_context"] == name:
        cfg["current_context"] = list(cfg["contexts"].keys())[0]
    save_cfg(cfg)
    return cfg


# ---------- start ----------
def main():
    # Jeśli dostępny CustomTkinter – uruchom modern UI
    if USE_CTK and ctk is not None:
        try:
            modern = ModernApp()
            modern.root.mainloop()
            return
        except Exception:
            # jeśli coś zawiedzie – spadamy do klasycznego UI
            pass

    root = TkBase()
    app = App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
